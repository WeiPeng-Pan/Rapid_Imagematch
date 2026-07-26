"""
物料图片匹配系统 - Tkinter 桌面应用
支持手动选择图片文件夹和物料清单 Excel，自动匹配并输出图片
"""

import os
import re
import sys
import math
import shutil
import threading
import queue
from pathlib import Path
from datetime import datetime

import tkinter as tk
from tkinter import ttk, filedialog, messagebox

try:
    import openpyxl
except ImportError:
    openpyxl = None

try:
    from PIL import Image, ImageTk
except ImportError:
    Image = None
    ImageTk = None

# 添加当前目录到路径，确保 matching 模块可导入
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from matching import clean_str, attention_score, build_idf_dict, build_candidate_text

# ===================== 配置 =====================
BASE_DIR = Path(__file__).parent
OUTPUT_DIR = BASE_DIR / "匹配结果输出"
THUMBNAIL_SIZE = (280, 280)
PREVIEW_SIZE = (400, 400)
DEFAULT_THRESHOLD = 50
DEFAULT_TOP_K = 3

# ===================== 列名映射 =====================
COLUMN_ALIASES = {
    "物料名称": ["物料名称", "名称", "物料名", "物资名称", "材料名称", "物品名称", "品名", "物料", "NAME"],
    "型号": ["型号", "规格型号", "型号规格", "规格", "MODEL", "TYPE", "规格型号(物料编码)"],
    "品牌": ["品牌", "商标", "BRAND", "厂家", "生产厂家", "制造商"],
    "参数": ["参数", "参数说明", "规格参数", "描述", "说明", "DESCRIPTION", "技术参数", "备注"],
}

# ===================== 工具函数 =====================
def sanitize_filename(s: str, max_len: int = 60) -> str:
    """清理字符串，使其适合作为文件名"""
    if not s:
        return ""
    # 替换不允许的文件名字符
    s = re.sub(r'[<>:"/\\|?*]', "_", s)
    s = re.sub(r'[\s]+', "_", s)
    s = re.sub(r'[_]+', "_", s)
    s = s.strip("_. ")
    if len(s) > max_len:
        s = s[:max_len]
    return s


def auto_detect_columns(headers):
    """
    自动识别 Excel 列名
    返回 { "物料名称": col_index, "型号": col_index, "品牌": col_index, "参数": col_index }
    未识别的列返回 None
    """
    detected = {"物料名称": None, "型号": None, "品牌": None, "参数": None}
    header_strs = [str(h).strip() if h else "" for h in headers]

    for key, aliases in COLUMN_ALIASES.items():
        for i, h in enumerate(header_strs):
            # 完全匹配或包含匹配（针对中文列名）
            h_clean = h.upper().replace(" ", "").replace("　", "")
            for alias in aliases:
                a_clean = alias.upper().replace(" ", "").replace("　", "")
                if h_clean == a_clean or h_clean == a_clean.replace("-", ""):
                    detected[key] = i
                    break
                # 宽松匹配：任意别名是表头的子串
                if len(alias) >= 4 and alias in h:
                    detected[key] = i
                    break
            if detected[key] is not None:
                break

    return detected


def generate_template_excel(filepath):
    """生成物料清单模板 Excel"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "物料清单"

    headers = ["物料名称", "型号", "品牌", "参数"]
    ws.append(headers)

    # 添加示例行
    ws.append(["齿轮箱弹性支撑", "14/002/003", "ESM", "48763"])
    ws.append(["高速刹车片磨损传感器", "490-3711-804", "SVENDBORG", ""])
    ws.append(["工业以太网交换机", "6-G20-M16-4POE", "运达", ""])

    # 设置列宽
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 30

    # 冻结首行
    ws.freeze_panes = "A2"

    wb.save(filepath)
    return filepath


# ===================== 数据加载器 =====================
class DataLoader:
    """数据加载模块"""

    @staticmethod
    def load_images(image_dir):
        """扫描图片目录，返回图片信息列表"""
        img_dir = Path(image_dir)
        if not img_dir.exists():
            raise FileNotFoundError(f"图片目录不存在: {img_dir}")

        images = []
        for f in sorted(img_dir.iterdir()):
            if f.suffix.lower() in (".jpg", ".jpeg", ".png", ".webp", ".gif", ".bmp"):
                stem = f.stem
                clean_stem = re.sub(r"[-_]\d+$", "", stem)
                images.append({
                    "path": str(f),
                    "filename": f.name,
                    "stem": stem,
                    "clean_stem": clean_stem,
                    "ext": f.suffix.lower(),
                })
        return images

    @staticmethod
    def load_excel(filepath):
        """
        加载任意 Excel 文件，自动识别列
        返回 (data_list, detected_columns)
        """
        filepath = Path(filepath)
        if not filepath.exists():
            raise FileNotFoundError(f"文件不存在: {filepath}")

        wb = openpyxl.load_workbook(filepath, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))

        if len(rows) < 2:
            wb.close()
            return [], {}

        headers = [str(h).strip() if h else "" for h in rows[0]]
        detected = auto_detect_columns(headers)

        # 检查是否至少识别了物料名称
        if detected["物料名称"] is None:
            wb.close()
            raise ValueError(
                f"未能识别「物料名称」列！\n检测到的表头: {headers}\n\n"
                f"支持的列名: {', '.join(COLUMN_ALIASES['物料名称'])}"
            )

        data = []
        for row in rows[1:]:
            name_idx = detected["物料名称"]
            if name_idx is not None and row[name_idx] and str(row[name_idx]).strip():
                item = {"物料名称": str(row[name_idx]).strip()}

                # 型号
                if detected["型号"] is not None:
                    val = row[detected["型号"]]
                    item["型号"] = str(val).strip() if val and str(val).strip() not in ("None", "") else ""
                else:
                    item["型号"] = ""

                # 品牌
                if detected["品牌"] is not None:
                    val = row[detected["品牌"]]
                    item["品牌"] = str(val).strip() if val and str(val).strip() not in ("None", "") else ""
                else:
                    item["品牌"] = ""

                # 参数
                if detected["参数"] is not None:
                    val = row[detected["参数"]]
                    item["参数"] = str(val).strip() if val and str(val).strip() not in ("None", "") else ""
                else:
                    item["参数"] = ""

                data.append(item)

        wb.close()
        return data, detected


# ===================== 匹配引擎 =====================
class MatcherEngine:
    """匹配引擎 - 封装注意力算法匹配逻辑"""

    def __init__(self, images, threshold=DEFAULT_THRESHOLD, top_k=DEFAULT_TOP_K):
        self.images = images
        self.threshold = threshold
        self.top_k = top_k
        self.idf_dict = None
        self._build_index()

    def _build_index(self):
        """构建 IDF 词典和候选文本列表"""
        candidate_texts = [clean_str(img["clean_stem"]) for img in self.images]
        self.idf_dict = build_idf_dict(candidate_texts)
        self.candidate_texts = candidate_texts

    def set_params(self, threshold=None, top_k=None):
        if threshold is not None:
            self.threshold = threshold
        if top_k is not None:
            self.top_k = top_k

    def match(self, item):
        """
        对单个物项执行匹配
        返回 [(图片信息, 分数), ...] 按分数降序排列
        """
        query = build_candidate_text(
            item["物料名称"], item.get("型号", ""), item.get("品牌", ""), item.get("参数", "")
        )
        if not query.strip():
            return []

        results = []
        for idx, img in enumerate(self.images):
            candidate = self.candidate_texts[idx]
            score = attention_score(query, candidate, self.idf_dict)
            if score >= self.threshold:
                results.append((img, round(score, 1)))

        results.sort(key=lambda x: x[1], reverse=True)
        return results[: self.top_k]

    def match_all(self, owner_data, progress_callback=None):
        """
        对全部物项执行匹配
        返回 [{item, matches, best_match, best_score}, ...]
        """
        results = []
        total = len(owner_data)

        for i, item in enumerate(owner_data):
            matches = self.match(item)
            best = matches[0] if matches else (None, 0)
            results.append({
                "item": item,
                "matches": matches,
                "best_img": best[0],
                "best_score": best[1],
                "matched": len(matches) > 0,
            })
            if progress_callback:
                progress_callback(i + 1, total, item["物料名称"])

        return results


# ===================== 导出器 =====================
class Exporter:
    """导出模块 - 支持按序号-物料名称-型号 重命名"""

    @staticmethod
    def export_all(results, export_dir, threshold_used, image_dir_src):
        """导出结果到指定目录"""
        export_dir = Path(export_dir)
        export_dir.mkdir(parents=True, exist_ok=True)
        img_dir = export_dir / "图片"
        img_dir.mkdir(parents=True, exist_ok=True)

        # 1. 复制图片并按「序号-物料名称-型号」重命名
        copied_files = []
        used_names = set()

        for i, res in enumerate(results, 1):
            item = res["item"]
            matches = res["matches"]

            for rank, (img, score) in enumerate(matches):
                # 构建新文件名
                seq = f"{i:02d}"
                name_part = sanitize_filename(item["物料名称"])
                model_part = sanitize_filename(item.get("型号", ""))
                if model_part:
                    new_name = f"{seq}-{name_part}-{model_part}{img['ext']}"
                else:
                    new_name = f"{seq}-{name_part}{img['ext']}"

                # 避免重名
                if new_name in used_names:
                    base = new_name[: new_name.rfind(".")]
                    new_name = f"{base}_v{rank}{img['ext']}"
                used_names.add(new_name)

                try:
                    shutil.copy2(img["path"], img_dir / new_name)
                    copied_files.append({
                        "original": img["filename"],
                        "renamed": new_name,
                        "score": score,
                        "item_idx": i,
                    })
                except Exception:
                    pass

        # 2. 生成 Excel 汇总
        excel_rows = []
        for i, res in enumerate(results, 1):
            item = res["item"]
            matches = res["matches"]
            if matches:
                for rank, (img, score) in enumerate(matches):
                    # 找对应的重命名文件
                    renamed = ""
                    for cf in copied_files:
                        if cf["original"] == img["filename"] and cf["item_idx"] == i:
                            renamed = cf["renamed"]
                            break
                    excel_rows.append({
                        "序号": i,
                        "物料名称": item["物料名称"],
                        "型号": item.get("型号", ""),
                        "品牌": item.get("品牌", ""),
                        "参数": item.get("参数", ""),
                        "匹配分数": score,
                        "原图片名": img["filename"],
                        "输出图片名": renamed,
                    })
            else:
                excel_rows.append({
                    "序号": i,
                    "物料名称": item["物料名称"],
                    "型号": item.get("型号", ""),
                    "品牌": item.get("品牌", ""),
                    "参数": item.get("参数", ""),
                    "匹配分数": "无",
                    "原图片名": "",
                    "输出图片名": "未匹配到",
                })

        excel_path = export_dir / "匹配结果汇总.xlsx"
        if openpyxl:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "匹配结果"

            headers = ["序号", "物料名称", "型号", "品牌", "参数", "匹配分数", "原图片名", "输出图片名"]
            ws.append(headers)
            for row_data in excel_rows:
                ws.append([row_data.get(h, "") for h in headers])

            for col_idx, _ in enumerate(headers, 1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = max(
                    len(headers[col_idx - 1]) * 3, 14
                )
            wb.save(excel_path)

        # 3. 生成 HTML 预览
        html_parts = [
            "<!DOCTYPE html><html><head><meta charset='utf-8'>",
            "<title>物料图片匹配结果</title>",
            "<style>",
            "body { font-family: -apple-system, sans-serif; max-width: 1200px; margin: 0 auto; padding: 20px; background: #f5f5f5; }",
            "h1 { color: #333; }",
            ".summary { margin: 16px 0; padding: 12px 20px; background: #fff; border-radius: 8px; box-shadow: 0 1px 3px rgba(0,0,0,.1); }",
            ".item { background: #fff; margin: 12px 0; padding: 16px 20px; border-radius: 8px; box-shadow: 0 1px 3px rgba(0,0,0,.1); }",
            ".item h3 { margin: 0 0 8px 0; color: #333; }",
            ".item .meta { color: #666; font-size: 14px; margin-bottom: 8px; }",
            ".item .meta span { margin-right: 16px; }",
            ".images { display: flex; flex-wrap: wrap; gap: 12px; margin-top: 8px; }",
            ".images .card { text-align: center; background: #fafafa; padding: 8px; border-radius: 6px; }",
            ".images .card img { max-width: 200px; max-height: 200px; border-radius: 4px; }",
            ".images .card .score { font-size: 13px; color: #e67e22; font-weight: bold; margin-top: 4px; }",
            ".no-match { color: #e74c3c; font-weight: bold; }",
            ".matched { border-left: 4px solid #27ae60; }",
            ".unmatched { border-left: 4px solid #e74c3c; }",
            "</style></head><body>",
            "<h1>物料图片匹配结果</h1>",
            f"<div class='summary'>",
            f"<p>生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>",
            f"<p>图片来源: {image_dir_src}</p>",
            f"<p>匹配阈值: {threshold_used}</p>",
            f"<p>总物料: {len(results)} 项 | 匹配成功: {sum(1 for r in results if r['matched'])} 项</p>",
            "</div>",
        ]

        for i, res in enumerate(results, 1):
            cls = "item matched" if res["matched"] else "item unmatched"
            item = res["item"]
            html_parts.append(f"<div class='{cls}'>")
            html_parts.append(f"<h3>#{i} {item['物料名称']}</h3>")
            html_parts.append("<div class='meta'>")
            if item.get("型号"):
                html_parts.append(f"<span>型号: {item['型号']}</span>")
            if item.get("品牌"):
                html_parts.append(f"<span>品牌: {item['品牌']}</span>")
            if item.get("参数"):
                html_parts.append(f"<span>参数: {item['参数']}</span>")
            html_parts.append("</div>")

            if res["matches"]:
                html_parts.append("<div class='images'>")
                for img, score in res["matches"]:
                    # 找对应的重命名文件
                    renamed = img["filename"]
                    for cf in copied_files:
                        if cf["original"] == img["filename"] and cf["item_idx"] == i:
                            renamed = cf["renamed"]
                            break
                    html_parts.append("<div class='card'>")
                    html_parts.append(f"<img src='图片/{renamed}' alt='{renamed}'>")
                    html_parts.append(f"<div class='score'>匹配度: {score}分</div>")
                    html_parts.append(f"<div style='font-size:11px;color:#999'>→ {renamed}</div>")
                    html_parts.append("</div>")
                html_parts.append("</div>")
            else:
                html_parts.append("<div class='no-match'>❌ 未匹配到图片</div>")
            html_parts.append("</div>")

        html_parts.append("</body></html>")
        html_path = export_dir / "预览.html"
        with open(html_path, "w", encoding="utf-8") as f:
            f.write("\n".join(html_parts))

        stats = {
            "total": len(results),
            "matched": sum(1 for r in results if r["matched"]),
            "images_copied": len(copied_files),
            "excel_path": str(excel_path),
            "html_path": str(html_path),
            "img_dir": str(img_dir),
        }
        return stats


# ===================== 图片预览弹窗 =====================
class ImagePreviewDialog(tk.Toplevel):
    def __init__(self, parent, img_path, title=""):
        super().__init__(parent)
        self.title(title or os.path.basename(img_path))
        self.geometry("800x700")

        try:
            img = Image.open(img_path)
            display_w, display_h = 750, 600
            img_w, img_h = img.size
            ratio = min(display_w / img_w, display_h / img_h)
            new_size = (int(img_w * ratio), int(img_h * ratio))
            img_resized = img.resize(new_size, Image.LANCZOS)

            photo = ImageTk.PhotoImage(img_resized)
            self.photo = photo

            label = ttk.Label(self, image=photo)
            label.pack(padx=20, pady=20)

            info = f"{os.path.basename(img_path)} | 原始尺寸: {img_w}x{img_h}"
            ttk.Label(self, text=info).pack()

        except Exception as e:
            ttk.Label(self, text=f"无法加载图片: {e}").pack(padx=20, pady=20)

        ttk.Button(self, text="关闭", command=self.destroy).pack(pady=10)


# ===================== 设置对话框 =====================
class SettingsDialog(tk.Toplevel):
    def __init__(self, parent, current_threshold, current_top_k, on_save):
        super().__init__(parent)
        self.title("匹配设置")
        self.geometry("350x200")
        self.resizable(False, False)
        self.on_save = on_save

        frame = ttk.Frame(self, padding=20)
        frame.pack(fill="both", expand=True)

        ttk.Label(frame, text="匹配阈值 (0-100):").grid(row=0, column=0, sticky="w", pady=5)
        self.threshold_var = tk.IntVar(value=current_threshold)
        ttk.Spinbox(frame, from_=10, to=95, textvariable=self.threshold_var, width=10).grid(
            row=0, column=1, pady=5, sticky="w"
        )

        ttk.Label(frame, text="每项展示图片数:").grid(row=1, column=0, sticky="w", pady=5)
        self.top_k_var = tk.IntVar(value=current_top_k)
        ttk.Spinbox(frame, from_=1, to=5, textvariable=self.top_k_var, width=10).grid(
            row=1, column=1, pady=5, sticky="w"
        )

        ttk.Label(frame, text="阈值越高匹配越精确\n阈值越低匹配越多", foreground="gray").grid(
            row=2, column=0, columnspan=2, pady=5
        )

        btn_frame = ttk.Frame(frame)
        btn_frame.grid(row=3, column=0, columnspan=2, pady=15)
        ttk.Button(btn_frame, text="保存", command=self._save).pack(side="left", padx=5)
        ttk.Button(btn_frame, text="取消", command=self.destroy).pack(side="left", padx=5)

    def _save(self):
        self.on_save(self.threshold_var.get(), self.top_k_var.get())
        self.destroy()


# ===================== 主应用 =====================
class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("物料图片匹配系统 v2.0")
        self.geometry("1200x750")
        self.minsize(1000, 650)

        # 数据状态
        self.owner_data = []
        self.images = []
        self.matcher = None
        self.match_results = []
        self.current_photo = None

        # 数据源路径
        self.image_dir_path = None
        self.excel_path = None

        # 参数
        self.threshold = DEFAULT_THRESHOLD
        self.top_k = DEFAULT_TOP_K

        # 样式
        self.style = ttk.Style()
        self.style.theme_use("clam")
        self._configure_styles()

        # 构建UI
        self._build_ui()

    def _configure_styles(self):
        self.style.configure("Success.TLabel", foreground="#27ae60", font=("", 11, "bold"))
        self.style.configure("Fail.TLabel", foreground="#e74c3c", font=("", 11, "bold"))
        self.style.configure("Score.TLabel", foreground="#e67e22", font=("", 14, "bold"))
        self.style.configure("Header.TLabel", font=("", 11, "bold"))
        self.style.configure("Title.TLabel", font=("", 18, "bold"))
        self.style.configure("SubTitle.TLabel", font=("", 12), foreground="#555")
        self.style.configure("Action.TButton", font=("", 11), padding=(16, 8))
        self.style.configure("Big.TButton", font=("", 13), padding=(20, 12))

    # ===================== UI 构建 =====================

    def _build_ui(self):
        """构建所有 UI 组件"""
        # 主容器
        self.container = ttk.Frame(self)
        self.container.pack(fill="both", expand=True)

        # 两个页面：欢迎页 + 主页面
        self.welcome_frame = ttk.Frame(self.container)
        self.main_frame = ttk.Frame(self.container)

        self._build_welcome_page()
        self._build_main_page()

        # 默认显示欢迎页
        self._show_welcome()

    def _build_welcome_page(self):
        """构建欢迎/选择页面"""
        f = self.welcome_frame

        # 居中内容
        center = ttk.Frame(f)
        center.place(relx=0.5, rely=0.45, anchor="center")

        # 标题
        ttk.Label(center, text="物料图片匹配系统", style="Title.TLabel").pack(pady=(0, 5))
        ttk.Label(center, text="选择图片文件夹和物料清单，自动匹配输出图片", style="SubTitle.TLabel").pack(pady=(0, 30))

        # --- 数据源选择区域 ---
        source_frame = ttk.LabelFrame(center, text="数据源选择", padding=(25, 15))
        source_frame.pack(pady=10)

        # 图片文件夹
        img_frame = ttk.Frame(source_frame)
        img_frame.pack(fill="x", pady=8)
        self.btn_select_img = ttk.Button(
            img_frame, text="📂  选择图片文件夹", command=self._select_image_folder, style="Big.TButton"
        )
        self.btn_select_img.pack(side="left")
        self.img_path_label = ttk.Label(img_frame, text="未选择", foreground="gray", wraplength=300)
        self.img_path_label.pack(side="left", padx=(15, 0))
        self.img_count_label = ttk.Label(img_frame, text="")
        self.img_count_label.pack(side="left", padx=(5, 0))

        # Excel 物料清单
        excel_frame = ttk.Frame(source_frame)
        excel_frame.pack(fill="x", pady=8)
        self.btn_select_excel = ttk.Button(
            excel_frame, text="📄  加载物料清单", command=self._select_excel_file, style="Big.TButton"
        )
        self.btn_select_excel.pack(side="left")
        self.excel_path_label = ttk.Label(excel_frame, text="未选择", foreground="gray", wraplength=300)
        self.excel_path_label.pack(side="left", padx=(15, 0))
        self.excel_count_label = ttk.Label(excel_frame, text="")
        self.excel_count_label.pack(side="left", padx=(5, 0))

        # 检测到的列名
        self.cols_label = ttk.Label(source_frame, text="", foreground="#666")
        self.cols_label.pack(fill="x", pady=(8, 0))

        # --- 操作按钮 ---
        btn_frame = ttk.Frame(center)
        btn_frame.pack(pady=(30, 10))

        self.btn_start = ttk.Button(
            btn_frame, text="🚀  开始匹配", command=self._show_main_after_load,
            style="Action.TButton", state="disabled"
        )
        self.btn_start.pack(side="left", padx=5)

        self.btn_template = ttk.Button(
            btn_frame, text="📋  下载模板", command=self._download_template, style="Action.TButton"
        )
        self.btn_template.pack(side="left", padx=5)

        # 状态提示
        self.welcome_status = ttk.Label(center, text="请选择图片文件夹和物料清单 Excel", foreground="gray")
        self.welcome_status.pack(pady=(15, 0))

        # 底部说明
        note_frame = ttk.Frame(center)
        note_frame.pack(pady=(25, 0))
        ttk.Label(note_frame, text="💡 物料清单需包含「物料名称」列，支持 Excel (.xlsx)", foreground="#999").pack()

    def _build_main_page(self):
        """构建主页面（三面板布局）"""
        f = self.main_frame

        # --- 顶部: 工具栏 ---
        toolbar = ttk.Frame(f, padding=(15, 10, 15, 5))
        toolbar.pack(fill="x")

        ttk.Label(toolbar, text="物料图片匹配系统", style="Header.TLabel").pack(side="left")

        self.btn_back = ttk.Button(toolbar, text="← 返回选择", command=self._show_welcome)
        self.btn_back.pack(side="left", padx=(15, 5))

        ttk.Separator(toolbar, orient="vertical").pack(side="left", fill="y", padx=5)

        self.btn_match_main = ttk.Button(
            toolbar, text="全量匹配", command=self._run_matching, style="Action.TButton", state="disabled"
        )
        self.btn_match_main.pack(side="left", padx=5)

        self.btn_export_main = ttk.Button(
            toolbar, text="导出结果", command=self._export_results, style="Action.TButton", state="disabled"
        )
        self.btn_export_main.pack(side="left", padx=5)

        self.btn_settings_main = ttk.Button(
            toolbar, text="设置", command=self._open_settings, style="Action.TButton"
        )
        self.btn_settings_main.pack(side="left", padx=5)

        # 数据源信息
        self.main_src_label = ttk.Label(toolbar, text="", foreground="gray", font=("", 9))
        self.main_src_label.pack(side="right")

        # 进度
        self.progress = ttk.Progressbar(toolbar, mode="determinate", length=150)
        self.progress.pack(side="right", padx=5)
        self.progress_label = ttk.Label(toolbar, text="", width=12)
        self.progress_label.pack(side="right", padx=5)

        # --- 三个面板 ---
        paned = ttk.PanedWindow(f, orient="horizontal")
        paned.pack(fill="both", expand=True, padx=15, pady=(5, 10))

        # 左: 物料清单
        left_frame = ttk.LabelFrame(paned, text="物料清单", padding=(5, 5))
        paned.add(left_frame, weight=1)
        self._build_material_list(left_frame)

        # 中: 物料详情 & 匹配结果
        center_frame = ttk.LabelFrame(paned, text="物料详情与匹配结果", padding=(10, 5))
        paned.add(center_frame, weight=2)
        self._build_detail_panel(center_frame)

        # 右: 图片预览
        right_frame = ttk.LabelFrame(paned, text="图片预览", padding=(5, 5))
        paned.add(right_frame, weight=2)
        self._build_preview_panel(right_frame)

        # --- 状态栏 ---
        self.status_var = tk.StringVar(value="就绪")
        status_bar = ttk.Label(f, textvariable=self.status_var, relief="sunken", anchor="w", padding=(10, 3))
        status_bar.pack(fill="x", side="bottom")

    def _build_material_list(self, parent):
        columns = ("idx", "name", "status")
        self.tree = ttk.Treeview(parent, columns=columns, show="headings", height=20, selectmode="browse")
        self.tree.heading("idx", text="#")
        self.tree.heading("name", text="物料名称")
        self.tree.heading("status", text="状态")
        self.tree.column("idx", width=40, anchor="center", minwidth=30)
        self.tree.column("name", width=180, minwidth=120)
        self.tree.column("status", width=50, anchor="center", minwidth=40)

        scrollbar = ttk.Scrollbar(parent, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=scrollbar.set)
        self.tree.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        self.tree.bind("<<TreeviewSelect>>", self._on_item_selected)

    def _build_detail_panel(self, parent):
        detail_frame = ttk.Frame(parent)
        detail_frame.pack(fill="x", pady=(0, 10))
        ttk.Label(detail_frame, text="物料信息", style="Header.TLabel").pack(anchor="w")

        info_grid = ttk.Frame(detail_frame)
        info_grid.pack(fill="x", pady=(5, 0))
        self.detail_labels = {}
        fields = [("名称:", "name"), ("型号:", "model"), ("品牌:", "brand"), ("参数:", "params")]
        for i, (label, key) in enumerate(fields):
            ttk.Label(info_grid, text=label, width=6, anchor="e").grid(
                row=i, column=0, sticky="e", padx=(0, 5), pady=2
            )
            lbl = ttk.Label(info_grid, text="-", anchor="w", wraplength=250)
            lbl.grid(row=i, column=1, sticky="w", pady=2)
            self.detail_labels[key] = lbl

        ttk.Separator(parent, orient="horizontal").pack(fill="x", pady=8)

        match_frame = ttk.Frame(parent)
        match_frame.pack(fill="both", expand=True)
        ttk.Label(match_frame, text="匹配结果", style="Header.TLabel").pack(anchor="w")
        self.match_status_label = ttk.Label(match_frame, text="等待匹配...")
        self.match_status_label.pack(anchor="w", pady=(5, 2))
        self.match_score_label = ttk.Label(match_frame, text="")
        self.match_score_label.pack(anchor="w", pady=(0, 5))
        self.match_files_frame = ttk.Frame(match_frame)
        self.match_files_frame.pack(fill="both", expand=True, pady=(5, 0))

    def _build_preview_panel(self, parent):
        self.preview_frame = ttk.Frame(parent)
        self.preview_frame.pack(fill="both", expand=True)
        self.preview_frame.pack_propagate(False)

        self.preview_placeholder = ttk.Label(
            self.preview_frame, text="选择物料后\n显示匹配图片", foreground="gray", justify="center"
        )
        self.preview_placeholder.pack(expand=True)

        self.preview_image_label = ttk.Label(self.preview_frame)
        self.preview_caption_label = ttk.Label(
            self.preview_frame, text="", foreground="gray", wraplength=350
        )
        self.preview_thumbs_frame = ttk.Frame(self.preview_frame)

        # --- 旋转状态 ---
        self.current_img_pil = None      # 原始 PIL Image
        self.current_img_path = None      # 当前图片路径
        self.current_filename = ""
        self.current_score = 0
        self.rotation_angle = 0           # 当前旋转角度 (0/90/180/270)

        # 旋转按钮工具条（初始隐藏）
        self.rotate_toolbar = ttk.Frame(self.preview_frame)
        self.btn_rotate_left = ttk.Button(
            self.rotate_toolbar, text="↺ 左旋", command=lambda: self._rotate(-90), width=8
        )
        self.btn_rotate_left.pack(side="left", padx=2)
        self.btn_rotate_right = ttk.Button(
            self.rotate_toolbar, text="↻ 右旋", command=lambda: self._rotate(90), width=8
        )
        self.btn_rotate_right.pack(side="left", padx=2)
        self.btn_rotate_reset = ttk.Button(
            self.rotate_toolbar, text="⟲ 复位", command=self._reset_rotation, width=8
        )
        self.btn_rotate_reset.pack(side="left", padx=2)
        self.rotate_angle_label = ttk.Label(self.rotate_toolbar, text="0°", width=4)
        self.rotate_angle_label.pack(side="left", padx=(8, 0))

    # ===================== 页面切换 =====================

    def _show_welcome(self):
        self.main_frame.pack_forget()
        self.welcome_frame.pack(fill="both", expand=True)
        self.welcome_status.configure(text="请选择图片文件夹和物料清单 Excel")

    def _show_main_after_load(self):
        """数据加载完成后，切换到主页面"""
        if not self.images or not self.owner_data:
            self._set_status("请先选择图片文件夹和物料清单")
            return

        self.welcome_frame.pack_forget()
        self.main_frame.pack(fill="both", expand=True)

        # 更新数据源信息
        src_text = f"📁 {os.path.basename(str(self.image_dir_path))} | 📄 {os.path.basename(str(self.excel_path))}"
        self.main_src_label.configure(text=src_text)

        # 刷新物料列表
        self._refresh_material_list()

        self.btn_match_main.configure(state="normal")
        self._set_status(
            f"已加载: {len(self.owner_data)} 项物料 | {len(self.images)} 张图片 | 点击「全量匹配」开始"
        )

    # ===================== 数据加载 =====================

    def _select_image_folder(self):
        dir_path = filedialog.askdirectory(title="选择图片文件夹")
        if not dir_path:
            return

        try:
            self.images = DataLoader.load_images(dir_path)
            self.image_dir_path = Path(dir_path)
            self.img_path_label.configure(text=str(self.image_dir_path))
            self.img_count_label.configure(
                text=f"({len(self.images)} 张)", foreground="#27ae60"
            )
            self._check_ready()
        except Exception as e:
            messagebox.showerror("加载失败", str(e))

    def _select_excel_file(self):
        file_path = filedialog.askopenfilename(
            title="选择物料清单",
            filetypes=[("Excel 文件", "*.xlsx *.xls"), ("所有文件", "*.*")],
        )
        if not file_path:
            return

        try:
            self.owner_data, detected_cols = DataLoader.load_excel(file_path)
            self.excel_path = Path(file_path)
            self.excel_path_label.configure(text=str(self.excel_path))
            self.excel_count_label.configure(
                text=f"({len(self.owner_data)} 项)", foreground="#27ae60"
            )

            # 显示识别的列名
            col_text = "识别列: "
            for key, idx in detected_cols.items():
                if idx is not None:
                    col_text += f"「{key}」✓  "
                else:
                    col_text += f"「{key}」✗  "
            self.cols_label.configure(text=col_text)

            self._check_ready()
        except ValueError as e:
            messagebox.showerror("列名识别失败", str(e))
        except Exception as e:
            messagebox.showerror("加载失败", str(e))

    def _check_ready(self):
        """检查是否已选择图片和Excel，启用开始按钮"""
        if self.images and self.owner_data:
            self.btn_start.configure(state="normal")
            self.welcome_status.configure(
                text=f"✅ 就绪！{len(self.owner_data)} 项物料, {len(self.images)} 张图片",
                foreground="#27ae60",
            )
        else:
            self.btn_start.configure(state="disabled")

    def _download_template(self):
        """下载模板 Excel"""
        save_path = filedialog.asksaveasfilename(
            title="保存模板",
            defaultextension=".xlsx",
            filetypes=[("Excel 文件", "*.xlsx")],
            initialfile="物料清单模板.xlsx",
        )
        if not save_path:
            return

        try:
            generate_template_excel(save_path)
            messagebox.showinfo(
                "模板已生成",
                f"✅ 模板已保存到:\n{save_path}\n\n"
                f"包含列: 物料名称、型号、品牌、参数\n"
                f"请填写数据后，在应用中选择此文件即可",
            )
        except Exception as e:
            messagebox.showerror("生成失败", str(e))

    def _refresh_material_list(self):
        for item in self.tree.get_children():
            self.tree.delete(item)
        for i, item in enumerate(self.owner_data, 1):
            name = item["物料名称"]
            name_short = name if len(name) <= 12 else name[:11] + "…"
            self.tree.insert("", "end", values=(i, name_short, "⏳"), iid=str(i))

    # ===================== 主页面交互 =====================

    def _on_item_selected(self, event):
        selection = self.tree.selection()
        if not selection:
            return
        idx = int(selection[0]) - 1
        if idx < 0 or idx >= len(self.owner_data):
            return

        item = self.owner_data[idx]
        self.detail_labels["name"].configure(text=item["物料名称"])
        self.detail_labels["model"].configure(text=item.get("型号", "-") or "-")
        self.detail_labels["brand"].configure(text=item.get("品牌", "-") or "-")
        self.detail_labels["params"].configure(text=item.get("参数", "-") or "-")

        for w in self.match_files_frame.winfo_children():
            w.destroy()

        if self.match_results and idx < len(self.match_results):
            res = self.match_results[idx]
            self._display_match_result(res)
        else:
            self.match_status_label.configure(text="尚未匹配")
            self.match_score_label.configure(text="")
            self._show_placeholder()

    def _display_match_result(self, result):
        item = result["item"]
        matches = result["matches"]

        if matches:
            best_img, best_score = matches[0]
            self.match_status_label.configure(text="✅ 已匹配", style="Success.TLabel")
            self.match_score_label.configure(text=f"最佳匹配分数: {best_score} 分", style="Score.TLabel")

            for img, score in matches:
                btn = ttk.Button(
                    self.match_files_frame,
                    text=f"  {img['filename'][:25]}… ({score}分)",
                    command=lambda p=img["path"], f=img["filename"], s=score: self._show_image(p, f, s),
                )
                btn.pack(fill="x", pady=1)

            self._show_image(best_img["path"], best_img["filename"], best_score)
        else:
            self.match_status_label.configure(text="❌ 未匹配到图片", style="Fail.TLabel")
            self.match_score_label.configure(text="")
            no_match_lbl = ttk.Label(
                self.match_files_frame,
                text="未找到匹配的图片文件\n请尝试降低匹配阈值",
                foreground="gray",
                justify="center",
            )
            no_match_lbl.pack(expand=True, pady=20)
            self._show_placeholder()

    def _show_image(self, img_path, filename, score):
        try:
            # 存储当前图片信息（用于旋转）
            self.current_img_pil = Image.open(img_path)
            self.current_img_path = img_path
            self.current_filename = filename
            self.current_score = score
            self.rotation_angle = 0

            self._render_preview()

            self.preview_placeholder.pack_forget()
            self.preview_thumbs_frame.pack_forget()
            self.preview_image_label.pack(expand=True, pady=(15, 5))
            self.preview_caption_label.pack(pady=(0, 5))

            # 显示旋转工具条
            self.rotate_toolbar.pack(pady=(0, 5))
            self.rotate_angle_label.configure(text="0°")

            self.preview_image_label.bind(
                "<Double-Button-1>",
                lambda e, p=img_path, f=filename: ImagePreviewDialog(self, p, f),
            )
            self.preview_image_label.configure(cursor="hand2")

        except Exception as e:
            self._show_placeholder(f"无法加载图片:\n{e}")

    def _render_preview(self):
        """根据当前旋转角度渲染图片"""
        if self.current_img_pil is None:
            return
        img = self.current_img_pil.copy()
        if self.rotation_angle != 0:
            img = img.rotate(self.rotation_angle, expand=True, resample=Image.BICUBIC)

        pw, ph = PREVIEW_SIZE
        iw, ih = img.size
        ratio = min(pw / iw, ph / ih)
        new_size = (int(iw * ratio), int(ih * ratio))
        img_resized = img.resize(new_size, Image.LANCZOS)
        photo = ImageTk.PhotoImage(img_resized)
        self.current_photo = photo
        self.preview_image_label.configure(image=photo)

        caption = f"{self.current_filename}  |  匹配分数: {self.current_score}"
        if self.rotation_angle != 0:
            caption += f"  |  已旋转 {self.rotation_angle}°"
        self.preview_caption_label.configure(text=caption)

    def _rotate(self, angle):
        """按角度旋转（正=顺时针，负=逆时针）"""
        self.rotation_angle = (self.rotation_angle + angle) % 360
        self.rotate_angle_label.configure(text=f"{self.rotation_angle}°")
        self._render_preview()

    def _reset_rotation(self):
        """复位旋转"""
        self.rotation_angle = 0
        self.rotate_angle_label.configure(text="0°")
        self._render_preview()

    def _show_placeholder(self, text=None):
        self.preview_image_label.pack_forget()
        self.preview_caption_label.pack_forget()
        self.preview_thumbs_frame.pack_forget()
        self.rotate_toolbar.pack_forget()  # 隐藏旋转工具条
        self.preview_placeholder.configure(text=text or "选择物料后\n显示匹配图片")
        self.preview_placeholder.pack(expand=True)

    # ===================== 匹配逻辑 =====================

    def _run_matching(self):
        if not self.owner_data or not self.matcher:
            if not self.matcher:
                self.matcher = MatcherEngine(self.images, self.threshold, self.top_k)
            if not self.owner_data:
                messagebox.showwarning("无数据", "请先加载物料清单")
                return

        self.btn_match_main.configure(state="disabled")
        self.btn_export_main.configure(state="disabled")
        self.match_results = []

        self._progress_queue = queue.Queue()

        def progress_callback(current, total, name):
            self._progress_queue.put((current, total, name))

        def matching_thread():
            try:
                results = self.matcher.match_all(self.owner_data, progress_callback)
                self._progress_queue.put(("DONE", results))
            except Exception as e:
                self._progress_queue.put(("ERROR", str(e)))

        thread = threading.Thread(target=matching_thread, daemon=True)
        thread.start()
        self._poll_progress()

    def _poll_progress(self):
        try:
            while True:
                msg = self._progress_queue.get_nowait()
                if msg[0] == "DONE":
                    self.match_results = msg[1]
                    self._on_matching_complete()
                    return
                elif msg[0] == "ERROR":
                    self._set_status(f"匹配失败: {msg[1]}")
                    messagebox.showerror("匹配失败", msg[1])
                    self.btn_match_main.configure(state="normal")
                    return
                else:
                    current, total, name = msg
                    self.progress["value"] = current / total * 100
                    self.progress_label.configure(text=f"{current}/{total}")
                    self._set_status(f"匹配中 ({current}/{total}): {name[:20]}…")
        except queue.Empty:
            pass
        self.after(50, self._poll_progress)

    def _on_matching_complete(self):
        self.progress["value"] = 100
        matched_count = sum(1 for r in self.match_results if r["matched"])

        self._set_status(
            f"匹配完成: {matched_count}/{len(self.match_results)} 项成功 "
            f"({matched_count / len(self.match_results) * 100:.0f}%)"
        )
        self.btn_match_main.configure(state="normal")
        self.btn_export_main.configure(state="normal")

        for i, res in enumerate(self.match_results, 1):
            status = "✅" if res["matched"] else "❌"
            item_id = str(i)
            if self.tree.exists(item_id):
                self.tree.set(item_id, "status", status)

        selection = self.tree.selection()
        if selection:
            self._on_item_selected(None)

        messagebox.showinfo(
            "匹配完成",
            f"✅ 全量匹配完成！\n\n总物料: {len(self.match_results)} 项\n"
            f"匹配成功: {matched_count} 项 ({matched_count / len(self.match_results) * 100:.0f}%)\n"
            f"未匹配: {len(self.match_results) - matched_count} 项\n\n"
            f"提示: 可点击「导出结果」保存到文件夹",
        )

    # ===================== 导出 =====================

    def _export_results(self):
        if not self.match_results:
            messagebox.showwarning("无结果", "请先执行全量匹配")
            return

        # 默认输出目录以物料清单文件名命名
        default_name = "匹配结果输出"
        if self.excel_path:
            default_name = f"匹配结果_{self.excel_path.stem}"

        dir_path = filedialog.askdirectory(title="选择导出目录", initialdir=str(OUTPUT_DIR))
        if not dir_path:
            return

        self._set_status("正在导出...")
        self.btn_export_main.configure(state="disabled")

        try:
            image_src = str(self.image_dir_path) if self.image_dir_path else ""
            stats = Exporter.export_all(self.match_results, dir_path, self.threshold, image_src)
            self._set_status(
                f"导出成功: {stats['matched']}/{stats['total']} 项匹配, "
                f"{stats['images_copied']} 张图片"
            )
            self.btn_export_main.configure(state="normal")

            # 图片已按「序号-物料名称-型号」重命名
            messagebox.showinfo(
                "导出完成",
                f"✅ 导出成功！\n\n保存路径: {dir_path}\n"
                f"📊 Excel汇总: 匹配结果汇总.xlsx\n"
                f"🌐 HTML预览: 预览.html\n"
                f"🖼️ 图片: {stats['images_copied']} 张 (已按序号-名称-型号重命名)\n\n"
                f"输出命名格式: 01-物料名称-型号.jpg",
            )

            import subprocess
            try:
                subprocess.run(["open", dir_path])
            except Exception:
                pass

        except Exception as e:
            self._set_status(f"导出失败: {e}")
            self.btn_export_main.configure(state="normal")
            messagebox.showerror("导出失败", str(e))

    # ===================== 设置 =====================

    def _open_settings(self):
        SettingsDialog(self, self.threshold, self.top_k, self._on_settings_save)

    def _on_settings_save(self, threshold, top_k):
        self.threshold = threshold
        self.top_k = top_k

        if self.matcher:
            self.matcher.set_params(threshold, top_k)

        self._set_status(f"设置已更新: 阈值={threshold}, 每项展示={top_k} 张")
        self.match_results = []

        for item in self.tree.get_children():
            self.tree.set(item, "status", "⏳")
        self.btn_export_main.configure(state="disabled")

    def _set_status(self, text):
        self.status_var.set(text)
        self.update_idletasks()


# ===================== 入口 =====================
def main():
    if not openpyxl:
        messagebox.showerror("缺少依赖", "请安装 openpyxl:\n  pip install openpyxl")
        return
    if not Image or not ImageTk:
        messagebox.showerror("缺少依赖", "请安装 Pillow:\n  pip install Pillow")
        return

    try:
        import jieba  # noqa
    except ImportError:
        messagebox.showwarning(
            "缺少 jieba",
            "建议安装 jieba 分词库以获得最佳匹配效果:\n  pip install jieba",
        )

    app = App()
    app.mainloop()


if __name__ == "__main__":
    main()
