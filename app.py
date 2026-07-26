"""
物料图片匹配系统 - Streamlit 应用
根据"业主需要上架物资"自动匹配并输出对应的图片
"""

import os
import re
import math
import shutil
from pathlib import Path

import streamlit as st
import openpyxl
import pandas as pd
from PIL import Image

from matching import (
    clean_str,
    attention_score,
    build_ngram_index,
    build_idf_dict,
    find_best_match,
    build_candidate_text,
)

# ===================== 配置 =====================
BASE_DIR = Path(__file__).parent

FILE_1_PATH = BASE_DIR / "1-原始命名.xlsx"
FILE_2_PATH = BASE_DIR / "2-业主需要上架物资.xlsx"
IMAGE_DIR = BASE_DIR / "图片程序测试" / "图片程序测试"

OUTPUT_DIR = BASE_DIR / "匹配结果输出"
THUMBNAIL_SIZE = (300, 300)

# ===================== 数据加载 =====================
@st.cache_data
def load_original_data():
    """加载原始命名文件"""
    wb = openpyxl.load_workbook(FILE_1_PATH, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    headers = rows[0]
    data = []
    for row in rows[1:]:
        if row[0] and str(row[0]).strip():
            data.append(
                {
                    "物料名称": str(row[0]).strip() if row[0] else "",
                    "型号": str(row[1]).strip() if row[1] and str(row[1]).strip() != "None" else "",
                    "参数": str(row[2]).strip() if row[2] and str(row[2]).strip() != "None" else "",
                    "品牌": str(row[3]).strip() if row[3] and str(row[3]).strip() != "None" else "",
                    "路径": str(row[4]).strip() if row[4] and str(row[4]).strip() != "None" else "",
                }
            )
    return data


@st.cache_data
def load_owner_data():
    """加载业主需要上架物资文件"""
    wb = openpyxl.load_workbook(FILE_2_PATH, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    headers = rows[0]
    data = []
    for row in rows[1:]:
        if row[0] and str(row[0]).strip():
            data.append(
                {
                    "物料名称": str(row[0]).strip() if row[0] else "",
                    "型号": str(row[1]).strip() if row[1] and str(row[1]).strip() != "None" else "",
                    "参数": str(row[2]).strip() if row[2] and str(row[2]).strip() != "None" else "",
                    "品牌": str(row[3]).strip() if row[3] and str(row[3]).strip() != "None" else "",
                }
            )
    return data


@st.cache_data
def load_images():
    """加载所有图片文件信息"""
    images = []
    if not IMAGE_DIR.exists():
        st.error(f"图片目录不存在: {IMAGE_DIR}")
        return images

    for f in sorted(IMAGE_DIR.iterdir()):
        if f.suffix.lower() in (".jpg", ".jpeg", ".png", ".webp", ".gif", ".bmp"):
            # 提取文件名（不含扩展名）作为匹配文本
            name_stem = f.stem
            # 去除末尾的 -1, -2, -3 等后缀用于匹配
            clean_stem = re.sub(r"[-_]\d+$", "", name_stem)
            images.append(
                {
                    "path": str(f),
                    "filename": f.name,
                    "stem": name_stem,
                    "clean_stem": clean_stem,
                    "ext": f.suffix.lower(),
                }
            )
    return images


@st.cache_data
def build_matching_index(original_data, images):
    """
    构建匹配索引：
    将所有候选文本（原始命名的物料+型号, 图片文件名等）整理成列表，
    并构建 IDF 词典和 n-gram 索引
    """
    candidates = []

    # 从原始数据构建候选
    for item in original_data:
        text = build_candidate_text(item["物料名称"], item["型号"], item["品牌"], item["参数"])
        candidates.append(
            {
                "type": "original",
                "text": text,
                "物料名称": item["物料名称"],
                "型号": item["型号"],
                "品牌": item["品牌"],
            }
        )

    # 从图片文件名构建候选
    for img in images:
        candidates.append(
            {
                "type": "image",
                "text": clean_str(img["clean_stem"]),
                "filename": img["filename"],
                "path": img["path"],
            }
        )

    # 提取纯文本列表
    candidate_texts = [c["text"] for c in candidates]

    # 构建 IDF
    idf_dict = build_idf_dict(candidate_texts)

    # 构建 n-gram 索引
    ngram_index = build_ngram_index(candidate_texts, n=3)

    return candidates, idf_dict, ngram_index


def match_item_to_images(
    item: dict,
    candidates: list,
    idf_dict: dict,
    ngram_index: dict,
    top_k: int = 3,
    threshold: float = 50.0,
):
    """
    将物项匹配到图片
    返回前 top_k 匹配结果
    """
    query = build_candidate_text(item["物料名称"], item["型号"], item["品牌"], item.get("参数", ""))

    results = []
    for i, cand in enumerate(candidates):
        score = attention_score(query, cand["text"], idf_dict)
        if score >= threshold and cand["type"] == "image":
            results.append(
                {
                    "score": round(score, 1),
                    "filename": cand["filename"],
                    "path": cand["path"],
                    "matched_text": cand["text"],
                }
            )

    # 按分数降序
    results.sort(key=lambda x: x["score"], reverse=True)
    return results[:top_k]


def match_item_to_original(
    item: dict,
    candidates: list,
    idf_dict: dict,
    ngram_index: dict,
    threshold: float = 50.0,
):
    """
    将物项匹配到原始命名中的最佳匹配项
    返回匹配到的原始项
    """
    query = build_candidate_text(item["物料名称"], item["型号"], item["品牌"], item.get("参数", ""))

    best_score = 0
    best_match = None

    for cand in candidates:
        if cand["type"] != "original":
            continue
        score = attention_score(query, cand["text"], idf_dict)
        if score > best_score:
            best_score = score
            best_match = cand

    if best_score >= threshold:
        return best_match, round(best_score, 1)
    return None, round(best_score, 1)


# ===================== 应用界面 =====================
st.set_page_config(
    page_title="物料图片匹配系统",
    page_icon="📦",
    layout="wide",
)

st.title("📦 物料图片自动匹配系统")
st.markdown(
    "根据「业主需要上架物资」自动匹配对应的产品图片"
)

# ---------- 侧边栏 ----------
with st.sidebar:
    st.header("⚙️ 设置")
    threshold = st.slider("匹配阈值", min_value=30, max_size=95, value=50, step=5, help="分数越高要求越精确")
    top_k = st.slider("每项展示图片数", min_value=1, max_value=5, value=3)

    st.divider()
    st.header("📂 文件状态")

    # 检查文件
    files_ok = True
    for fpath, label in [(FILE_1_PATH, "原始命名"), (FILE_2_PATH, "业主清单"), (IMAGE_DIR, "图片目录")]:
        exists = fpath.exists()
        st.markdown(f"{'✅' if exists else '❌'} **{label}**")
        if not exists:
            files_ok = False
            st.caption(str(fpath))

    if not files_ok:
        st.error("部分文件/目录缺失，请检查路径配置")

    st.divider()
    st.header("ℹ️ 说明")
    st.markdown(
        """
        - 匹配算法基于**注意力机制**，综合中文分词、字符n-gram、TF-IDF加权、位置衰减等
        - 匹配度分数 0-100，越高表示越匹配
        - 可通过阈值滑块调节匹配精度
        - 点击图片可查看原图
        """
    )

# ---------- 主内容 ----------
if not files_ok:
    st.stop()

# 加载数据
with st.spinner("正在加载数据..."):
    original_data = load_original_data()
    owner_data = load_owner_data()
    images = load_images()

st.success(f"✅ 原始物料: {len(original_data)} 项 | 业主清单: {len(owner_data)} 项 | 图片: {len(images)} 张")

# 构建匹配索引
with st.spinner("正在构建匹配索引..."):
    candidates, idf_dict, ngram_index = build_matching_index(original_data, images)

# ---------- 全量匹配按钮 ----------
col1, col2, col3 = st.columns([2, 1, 2])
with col1:
    run_all = st.button("🚀 执行全量匹配", type="primary", use_container_width=True)
with col3:
    export = st.button("📥 导出匹配结果", use_container_width=True)

# ---------- 匹配结果 ----------
if run_all:
    st.divider()
    st.header("📊 匹配结果")

    progress_bar = st.progress(0)
    status_text = st.empty()

    # 存储所有匹配结果
    all_results = []

    for idx, item in enumerate(owner_data):
        status_text.text(f"匹配中 ({idx+1}/{len(owner_data)}): {item['物料名称']}")
        progress_bar.progress((idx + 1) / len(owner_data))

        # 先找原始匹配
        orig_match, orig_score = match_item_to_original(
            item, candidates, idf_dict, ngram_index, threshold=threshold
        )

        # 再找图片匹配
        img_matches = match_item_to_images(
            item, candidates, idf_dict, ngram_index, top_k=top_k, threshold=threshold
        )

        all_results.append(
            {
                "index": idx + 1,
                "item": item,
                "orig_match": orig_match,
                "orig_score": orig_score,
                "img_matches": img_matches,
            }
        )

    progress_bar.empty()
    status_text.empty()

    # 显示结果
    # 统计
    matched_count = sum(1 for r in all_results if r["img_matches"])
    st.metric("匹配成功率", f"{matched_count}/{len(all_results)}",
              f"{matched_count/len(all_results)*100:.0f}%" if all_results else "0%")

    # 逐项显示
    for result in all_results:
        item = result["item"]
        orig_match = result["orig_match"]
        orig_score = result["orig_score"]
        img_matches = result["img_matches"]

        with st.container():
            cols = st.columns([1, 3, 4])

            # 序号
            with cols[0]:
                st.markdown(f"### #{result['index']}")

            # 物料信息
            with cols[1]:
                st.markdown(f"**{item['物料名称']}**")
                if item.get("型号"):
                    st.caption(f"型号: {item['型号']}")
                if item.get("品牌"):
                    st.caption(f"品牌: {item['品牌']}")
                if item.get("参数"):
                    st.caption(f"参数: {item['参数']}")

                # 原始匹配情况
                if orig_match:
                    st.markdown(f"📎 **原始匹配** (分数: {orig_score})")
                    st.caption(f"→ {orig_match['物料名称']} | {orig_match['型号']}")

            # 图片显示
            with cols[2]:
                if img_matches:
                    # 用列来展示多张图片
                    img_cols = st.columns(min(len(img_matches), 3))
                    for i, match in enumerate(img_matches):
                        col_idx = i % 3
                        with img_cols[col_idx]:
                            try:
                                img = Image.open(match["path"])
                                img.thumbnail(THUMBNAIL_SIZE)
                                st.image(img, caption=f"{match['score']}分", use_container_width=True)
                            except Exception as e:
                                st.error(f"无法加载图片: {match['filename']}")
                            st.caption(f"📄 {match['filename'][:30]}...")
                else:
                    st.warning("⚠️ 未找到匹配图片")

            st.divider()

    # 保存结果到 session state 用于导出
    st.session_state["all_results"] = all_results

elif "all_results" in st.session_state:
    # 显示缓存的匹配结果
    st.info("匹配结果已在会话中，可点击「导出匹配结果」保存到文件")

# ---------- 导出功能 ----------
if export or st.session_state.get("export_trigger"):
    if "all_results" not in st.session_state or not st.session_state["all_results"]:
        st.warning("请先执行全量匹配，再导出结果")
    else:
        # 导出为 Excel + 图片复制
        OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
        export_dir = OUTPUT_DIR / f"匹配结果_{pd.Timestamp.now().strftime('%Y%m%d_%H%M%S')}"
        export_dir.mkdir(parents=True, exist_ok=True)
        image_export_dir = export_dir / "图片"
        image_export_dir.mkdir(parents=True, exist_ok=True)

        # 准备导出数据
        export_rows = []
        for result in st.session_state["all_results"]:
            item = result["item"]
            orig_match = result["orig_match"]
            img_matches = result["img_matches"]

            if img_matches:
                for i, match in enumerate(img_matches):
                    # 复制图片
                    src = match["path"]
                    dst = image_export_dir / match["filename"]
                    try:
                        shutil.copy2(src, dst)
                    except Exception:
                        pass

                    export_rows.append(
                        {
                            "序号": result["index"],
                            "物料名称": item["物料名称"],
                            "型号": item.get("型号", ""),
                            "品牌": item.get("品牌", ""),
                            "参数": item.get("参数", ""),
                            "匹配分数": match["score"],
                            "匹配图片": match["filename"],
                            "原始匹配名称": orig_match["物料名称"] if orig_match else "",
                            "原始匹配型号": orig_match["型号"] if orig_match else "",
                            "原始匹配分数": result["orig_score"],
                        }
                    )
            else:
                export_rows.append(
                    {
                        "序号": result["index"],
                        "物料名称": item["物料名称"],
                        "型号": item.get("型号", ""),
                        "品牌": item.get("品牌", ""),
                        "参数": item.get("参数", ""),
                        "匹配分数": "",
                        "匹配图片": "未找到",
                        "原始匹配名称": orig_match["物料名称"] if orig_match else "",
                        "原始匹配型号": orig_match["型号"] if orig_match else "",
                        "原始匹配分数": result["orig_score"] if result["orig_score"] >= threshold else "",
                    }
                )

        # 写 Excel
        df = pd.DataFrame(export_rows)
        excel_path = export_dir / "匹配结果汇总.xlsx"
        df.to_excel(excel_path, index=False)

        # 写一个简单的 HTML 预览
        html_parts = [
            "<!DOCTYPE html><html><head><meta charset='utf-8'>",
            "<title>物料图片匹配结果</title>",
            "<style>",
            "body { font-family: sans-serif; max-width: 1200px; margin: 0 auto; padding: 20px; }",
            "table { border-collapse: collapse; width: 100%; }",
            "th, td { border: 1px solid #ddd; padding: 8px; text-align: left; }",
            "th { background-color: #f2f2f2; }",
            "img { max-width: 200px; max-height: 200px; }",
            ".matched { background-color: #e8f5e9; }",
            ".unmatched { background-color: #ffebee; }",
            "</style></head><body>",
            "<h1>物料图片匹配结果</h1>",
            f"<p>生成时间: {pd.Timestamp.now().strftime('%Y-%m-%d %H:%M:%S')}</p>",
            "<table><tr>",
            "<th>序号</th><th>物料名称</th><th>型号</th><th>品牌</th><th>匹配分数</th><th>匹配图片</th>",
            "</tr>",
        ]

        for row in export_rows:
            cls = "matched" if row["匹配分数"] else "unmatched"
            html_parts.append(f"<tr class='{cls}'>")
            html_parts.append(f"<td>{row['序号']}</td>")
            html_parts.append(f"<td>{row['物料名称']}</td>")
            html_parts.append(f"<td>{row['型号']}</td>")
            html_parts.append(f"<td>{row['品牌']}</td>")
            html_parts.append(f"<td>{row['匹配分数']}</td>")
            if row["匹配图片"] and row["匹配图片"] != "未找到":
                html_parts.append(
                    f"<td><img src='图片/{row['匹配图片']}'><br>{row['匹配图片']}</td>"
                )
            else:
                html_parts.append(f"<td>❌ {row['匹配图片']}</td>")
            html_parts.append("</tr>")

        html_parts.append("</table></body></html>")

        html_path = export_dir / "预览.html"
        with open(html_path, "w", encoding="utf-8") as f:
            f.write("\n".join(html_parts))

        st.success(f"✅ 导出成功！文件保存在: {export_dir}")
        st.info(f"📊 Excel汇总: {excel_path.name}")
        st.info(f"🖼️ 图片: {image_export_dir.name}/ ({sum(1 for r in export_rows if r['匹配图片'] and r['匹配图片'] != '未找到')} 张)")
        st.info(f"🌐 HTML预览: {html_path.name}")

        # 清理 session state
        if export:
            st.session_state["export_trigger"] = False

# ---------- 单项查询模式 ----------
st.divider()
st.header("🔍 单项查询")
st.markdown("手动输入物料信息，查询匹配的图片")

with st.form(key="single_query"):
    col1, col2, col3 = st.columns(3)
    with col1:
        q_name = st.text_input("物料名称")
    with col2:
        q_model = st.text_input("型号")
    with col3:
        q_brand = st.text_input("品牌")

    q_submit = st.form_submit_button("🔍 查询")

if q_submit and q_name:
    query_item = {"物料名称": q_name, "型号": q_model, "品牌": q_brand, "参数": ""}

    orig_match, orig_score = match_item_to_original(
        query_item, candidates, idf_dict, ngram_index, threshold=0
    )

    img_matches = match_item_to_images(
        query_item, candidates, idf_dict, ngram_index, top_k=5, threshold=0
    )

    col1, col2 = st.columns([1, 1])

    with col1:
        st.markdown("**原始匹配**")
        if orig_match:
            st.info(f"匹配: {orig_match['物料名称']} | {orig_match['型号']} (分数: {orig_score})")
        else:
            st.warning(f"无匹配 (最佳分数: {orig_score})")

    with col2:
        st.markdown("**图片匹配**")
        if img_matches:
            img_cols = st.columns(min(len(img_matches), 3))
            for i, match in enumerate(img_matches[:3]):
                with img_cols[i]:
                    try:
                        img = Image.open(match["path"])
                        img.thumbnail(THUMBNAIL_SIZE)
                        st.image(img, caption=f"{match['score']}分", use_container_width=True)
                    except Exception:
                        st.error("加载失败")
                    st.caption(f"{match['filename'][:30]}")
        else:
            st.warning("未找到匹配图片")

# ---------- 底部 ----------
st.divider()
st.caption("物料图片自动匹配系统 v1.0")
