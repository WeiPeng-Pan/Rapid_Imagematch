"""
物料图片匹配系统 - CustomTkinter 蓝白版
支持：图片缩放/旋转、蓝白UI、鼠标滚轮交互、手动选图
"""

import os
import re
import sys
import math
import shutil
import threading
import queue
from pathlib import Path
from datetime import datetime

import tkinter as tk
from tkinter import filedialog, messagebox

import customtkinter as ctk
from PIL import Image, ImageTk

if getattr(sys, 'frozen', False):
    sys.path.insert(0, sys._MEIPASS)
else:
    sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from matching import clean_str, attention_score, build_idf_dict, build_candidate_text

# ===================== 配置 =====================
ctk.set_appearance_mode("light")
ctk.set_default_color_theme("blue")

# 兼容 PyInstaller 打包模式（frozen）
if getattr(sys, 'frozen', False):
    BASE_DIR = Path(sys.executable).parent
else:
    BASE_DIR = Path(__file__).parent
OUTPUT_DIR = BASE_DIR / "匹配结果输出"

ZOOM_STEP = 0.15
ZOOM_MIN = 0.25
ZOOM_MAX = 4.0
ZOOM_FIT = -1  # 标记"适应窗口"

THUMBNAIL_SIZE = (260, 260)
THUMB_CARD_SIZE = (70, 70)  # 候选图片缩略图尺寸
DEFAULT_THRESHOLD = 50
DEFAULT_TOP_K = 3

COLUMN_ALIASES = {
    "物料名称": ["物料名称", "名称", "物料名", "物资名称", "材料名称", "物品名称", "品名", "物料", "NAME"],
    "型号": ["型号", "规格型号", "型号规格", "规格", "MODEL", "TYPE", "规格型号(物料编码)"],
    "品牌": ["品牌", "商标", "BRAND", "厂家", "生产厂家", "制造商"],
    "参数": ["参数", "参数说明", "规格参数", "描述", "说明", "DESCRIPTION", "技术参数", "备注"],
}

# ===================== 工具函数 =====================
def sanitize_filename(s: str, max_len: int = 60) -> str:
    if not s:
        return ""
    s = re.sub(r'[<>:"/\\|?*]', "_", s)
    s = re.sub(r'[\s]+', "_", s)
    s = re.sub(r'[_]+', "_", s)
    s = s.strip("_. ")
    return s[:max_len]

def auto_detect_columns(headers):
    detected = {"物料名称": None, "型号": None, "品牌": None, "参数": None}
    header_strs = [str(h).strip() if h else "" for h in headers]
    for key, aliases in COLUMN_ALIASES.items():
        for i, h in enumerate(header_strs):
            h_clean = h.upper().replace(" ", "").replace("　", "")
            for alias in aliases:
                a_clean = alias.upper().replace(" ", "").replace("　", "")
                if h_clean == a_clean or h_clean == a_clean.replace("-", ""):
                    detected[key] = i
                    break
                if len(alias) >= 4 and alias in h:
                    detected[key] = i
                    break
            if detected[key] is not None:
                break
    return detected

def generate_template_excel(filepath):
    wb = __import__("openpyxl").Workbook()
    ws = wb.active
    ws.title = "物料清单"
    ws.append(["物料名称", "型号", "品牌", "参数"])
    ws.append(["齿轮箱弹性支撑", "14/002/003", "ESM", "48763"])
    ws.append(["高速刹车片磨损传感器", "490-3711-804", "SVENDBORG", ""])
    ws.append(["工业以太网交换机", "6-G20-M16-4POE", "运达", ""])
    for col, w in [("A", 25), ("B", 30), ("C", 20), ("D", 30)]:
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"
    wb.save(filepath)
    return filepath

# ===================== 数据加载器 =====================
class DataLoader:
    @staticmethod
    def load_images(image_dir):
        img_dir = Path(image_dir)
        if not img_dir.exists():
            raise FileNotFoundError(f"图片目录不存在: {img_dir}")
        images = []
        # 递归扫描所有子目录（rglob 递归匹配所有文件）
        for f in sorted(img_dir.rglob("*")):
            if f.is_file() and f.suffix.lower() in (".jpg", ".jpeg", ".png", ".webp", ".gif", ".bmp"):
                stem = f.stem
                clean_stem = re.sub(r"[-_]\d+$", "", stem)
                images.append({
                    "path": str(f), "filename": f.name, "stem": stem,
                    "clean_stem": clean_stem, "ext": f.suffix.lower(),
                })
        return images

    @staticmethod
    def load_excel(filepath):
        import openpyxl
        filepath = Path(filepath)
        if not filepath.exists():
            raise FileNotFoundError(f"文件不存在: {filepath}")
        wb = openpyxl.load_workbook(filepath, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            wb.close()
            return [], {}
        headers = [str(h).strip() if h else "" for h in rows[0]]
        detected = auto_detect_columns(headers)
        if detected["物料名称"] is None:
            wb.close()
            raise ValueError(f"未能识别「物料名称」列！\n检测到的表头: {headers}")
        data = []
        for row in rows[1:]:
            ni = detected["物料名称"]
            if ni is not None and row[ni] and str(row[ni]).strip():
                item = {"物料名称": str(row[ni]).strip()}
                for key, idx_key in [("型号", "型号"), ("品牌", "品牌"), ("参数", "参数")]:
                    idx = detected.get(idx_key)
                    if idx is not None:
                        v = row[idx]
                        item[key] = str(v).strip() if v and str(v).strip() not in ("None", "") else ""
                    else:
                        item[key] = ""
                data.append(item)
        wb.close()
        return data, detected

# ===================== 匹配引擎 =====================
class MatcherEngine:
    def __init__(self, images, threshold=DEFAULT_THRESHOLD, top_k=DEFAULT_TOP_K):
        self.images = images
        self.threshold = threshold
        self.top_k = top_k
        self._build_index()

    def _build_index(self):
        self.candidate_texts = [clean_str(img["clean_stem"]) for img in self.images]
        self.idf_dict = build_idf_dict(self.candidate_texts)

    def set_params(self, threshold=None, top_k=None):
        if threshold is not None: self.threshold = threshold
        if top_k is not None: self.top_k = top_k

    def match(self, item):
        query = build_candidate_text(
            item["物料名称"], item.get("型号", ""), item.get("品牌", ""), item.get("参数", "")
        )
        if not query.strip():
            return []
        results = []
        for idx, img in enumerate(self.images):
            score = attention_score(query, self.candidate_texts[idx], self.idf_dict)
            if score >= self.threshold:
                results.append((img, round(score, 1)))
        results.sort(key=lambda x: x[1], reverse=True)
        return results[:self.top_k]

    def match_all(self, owner_data, progress_callback=None):
        results = []
        for i, item in enumerate(owner_data):
            matches = self.match(item)
            best = matches[0] if matches else (None, 0)
            results.append({
                "item": item, "matches": matches,
                "best_img": best[0], "best_score": best[1],
                "matched": len(matches) > 0,
            })
            if progress_callback:
                progress_callback(i + 1, len(owner_data), item["物料名称"])
        return results

# ===================== 导出器 =====================
class Exporter:
    @staticmethod
    def _pick_image(results, user_selections, item_idx):
        """根据用户选择或最佳匹配，返回 (img_info, score) 或 None"""
        if item_idx < 0 or item_idx >= len(results):
            return None
        res = results[item_idx]
        if not res["matches"]:
            return None

        # 用户手动选择了某张图
        if user_selections and item_idx in user_selections:
            sel_path = user_selections[item_idx]
            for img, score in res["matches"]:
                if img["path"] == sel_path:
                    return (img, score)

        # 默认取最佳匹配
        return res["matches"][0]

    @staticmethod
    def export_all(results, export_dir, threshold_used, image_dir_src, user_selections=None):
        import openpyxl
        export_dir = Path(export_dir)
        export_dir.mkdir(parents=True, exist_ok=True)
        img_dir = export_dir / "图片"
        img_dir.mkdir(parents=True, exist_ok=True)
        copied, used_names = [], set()

        # 只导出每项选中的那张图（最佳匹配或用户手动选择）
        for i in range(len(results)):
            picked = Exporter._pick_image(results, user_selections, i)
            if picked is None:
                continue
            img, score = picked
            item = results[i]["item"]
            seq = f"{i+1:02d}"
            name_part = sanitize_filename(item["物料名称"])
            model_part = sanitize_filename(item.get("型号", ""))
            new_name = f"{seq}-{name_part}-{model_part}{img['ext']}" if model_part else f"{seq}-{name_part}{img['ext']}"
            if new_name in used_names:
                base = new_name[:new_name.rfind(".")]
                new_name = f"{base}_v{seq}{img['ext']}"
            used_names.add(new_name)
            try:
                shutil.copy2(img["path"], img_dir / new_name)
                copied.append({"original": img["filename"], "renamed": new_name, "score": score, "item_idx": i + 1})
            except Exception:
                pass

        excel_rows = []
        for i in range(len(results)):
            item = results[i]["item"]
            picked = Exporter._pick_image(results, user_selections, i)
            if picked:
                img, score = picked
                renamed = next((c["renamed"] for c in copied if c["original"] == img["filename"] and c["item_idx"] == i + 1), img["filename"])
                excel_rows.append({"序号": i + 1, "物料名称": item["物料名称"], "型号": item.get("型号",""), "品牌": item.get("品牌",""), "参数": item.get("参数",""), "匹配分数": score, "输出图片名": renamed})
            else:
                excel_rows.append({"序号": i + 1, "物料名称": item["物料名称"], "型号": item.get("型号",""), "品牌": item.get("品牌",""), "参数": item.get("参数",""), "匹配分数": "无", "输出图片名": "未匹配到"})

        if excel_rows:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "匹配结果"
            headers = ["序号", "物料名称", "型号", "品牌", "参数", "匹配分数", "输出图片名"]
            ws.append(headers)
            for rd in excel_rows:
                ws.append([rd.get(h,"") for h in headers])
            wb.save(export_dir / "匹配结果汇总.xlsx")

        html_path = Exporter._gen_html(results, copied, export_dir, threshold_used, image_dir_src, user_selections)
        return {"total": len(results), "matched": sum(1 for r in results if r["matched"]), "images_copied": len(copied), "excel_path": str(export_dir / "匹配结果汇总.xlsx"), "html_path": str(html_path), "img_dir": str(img_dir)}

    @staticmethod
    def _gen_html(results, copied, export_dir, threshold_used, image_dir_src, user_selections=None):
        html = [
            "<!DOCTYPE html><html><head><meta charset='utf-8'><title>物料图片匹配结果</title>",
            "<style>",
            "body{font-family:-apple-system,sans-serif;max-width:1200px;margin:0 auto;padding:20px;background:#f5f5f5}",
            "h1{color:#333}.summary{margin:16px 0;padding:12px 20px;background:#fff;border-radius:8px;box-shadow:0 1px 3px rgba(0,0,0,.1)}",
            ".item{background:#fff;margin:12px 0;padding:16px 20px;border-radius:8px;box-shadow:0 1px 3px rgba(0,0,0,.1)}",
            ".item h3{margin:0 0 8px 0;color:#333}.item .meta{color:#666;font-size:14px;margin-bottom:8px}",
            ".images{display:flex;flex-wrap:wrap;gap:12px;margin-top:8px}",
            ".card{text-align:center;background:#fafafa;padding:8px;border-radius:6px}",
            ".card img{max-width:200px;max-height:200px;border-radius:4px}",
            ".card .score{font-size:13px;color:#e67e22;font-weight:bold;margin-top:4px}",
            ".no-match{color:#e74c3c;font-weight:bold}",
            ".matched{border-left:4px solid #27ae60}.unmatched{border-left:4px solid #e74c3c}",
            "</style></head><body><h1>物料图片匹配结果</h1>",
            f"<div class='summary'><p>生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>",
            f"<p>图片来源: {image_dir_src}</p><p>匹配阈值: {threshold_used}</p>",
            f"<p>总物料: {len(results)} 项 | 匹配成功: {sum(1 for r in results if r['matched'])} 项</p></div>"
        ]
        for i in range(len(results)):
            res = results[i]
            cls = "item matched" if res["matched"] else "item unmatched"
            item = res["item"]
            html.append(f"<div class='{cls}'><h3>#{i+1} {item['物料名称']}</h3><div class='meta'>")
            if item.get("型号"): html.append(f"<span>型号: {item['型号']}</span>")
            if item.get("品牌"): html.append(f"<span>品牌: {item['品牌']}</span>")
            if item.get("参数"): html.append(f"<span>参数: {item['参数']}</span>")
            html.append("</div>")
            picked = Exporter._pick_image(results, user_selections, i)
            if picked:
                img, score = picked
                renamed = next((c["renamed"] for c in copied if c["original"] == img["filename"] and c["item_idx"] == i + 1), img["filename"])
                html.append("<div class='images'>")
                html.append(f"<div class='card'><img src='图片/{renamed}' alt='{renamed}'><div class='score'>匹配度: {score}分</div><div style='font-size:11px;color:#999'>→ {renamed}</div></div>")
                html.append("</div>")
            else:
                html.append("<div class='no-match'>❌ 未匹配到图片</div>")
            html.append("</div>")
        html.append("</body></html>")
        html_path = export_dir / "预览.html"
        with open(html_path, "w", encoding="utf-8") as f:
            f.write("\n".join(html))
        return html_path

# ===================== 大图预览弹窗 =====================
class ImagePreviewWindow(ctk.CTkToplevel):
    def __init__(self, parent, img_path, title=""):
        super().__init__(parent)
        self.title(title or os.path.basename(img_path))
        self.geometry("900x800")
        self.minsize(600, 500)

        self.img_path = img_path
        self.pil_img = Image.open(img_path)
        self.zoom_level = 1.0

        # 主图区域
        self.image_label = ctk.CTkLabel(self, text="")
        self.image_label.pack(fill="both", expand=True, padx=15, pady=(15, 5))
        self._render()

        # 底部工具条
        toolbar = ctk.CTkFrame(self)
        toolbar.pack(fill="x", padx=15, pady=(5, 15))

        ctk.CTkButton(toolbar, text="−", width=35, command=lambda: self._zoom(ZOOM_STEP * -1)).pack(side="left", padx=2)
        self.zoom_label = ctk.CTkLabel(toolbar, text="100%", width=50)
        self.zoom_label.pack(side="left", padx=5)
        ctk.CTkButton(toolbar, text="+", width=35, command=lambda: self._zoom(ZOOM_STEP)).pack(side="left", padx=2)
        ctk.CTkButton(toolbar, text="⟲ 适应", width=70, command=self._fit).pack(side="left", padx=10)
        ctk.CTkButton(toolbar, text="关闭", command=self.destroy).pack(side="right")

        # 鼠标滚轮绑定
        self.image_label.bind("<MouseWheel>", self._on_mousewheel)
        self.image_label.focus_set()

        info = f"{os.path.basename(img_path)}  |  原始: {self.pil_img.width}×{self.pil_img.height}"
        ctk.CTkLabel(toolbar, text=info, text_color="gray").pack(side="right", padx=10)

    def _render(self):
        w, h = 750, 650
        iw, ih = self.pil_img.size
        r = min(w / iw, h / ih) * self.zoom_level
        ns = (max(50, int(iw * r)), max(50, int(ih * r)))
        img_r = self.pil_img.resize(ns, Image.LANCZOS)
        ctk_img = ctk.CTkImage(img_r, size=ns)
        self.image_label.configure(image=ctk_img, text="")
        self._ctk_img_ref = ctk_img
        self.zoom_label.configure(text=f"{int(self.zoom_level * 100)}%")

    def _zoom(self, delta):
        self.zoom_level = max(ZOOM_MIN, min(ZOOM_MAX, self.zoom_level + delta))
        self._render()

    def _fit(self):
        self.zoom_level = 1.0
        self._render()

    def _on_mousewheel(self, event):
        delta = ZOOM_STEP if event.delta > 0 else -ZOOM_STEP
        self._zoom(delta)

# ===================== 物料行组件 =====================
class MaterialRow(ctk.CTkFrame):
    """物料列表中的可点击行"""
    def __init__(self, parent, index, name, status="⏳", is_selected=False, command=None):
        super().__init__(parent, corner_radius=6, height=32, fg_color="transparent")
        self.index = index
        self.command = command
        self.is_selected = is_selected
        self.pack(fill="x", pady=1)

        self.grid_columnconfigure(1, weight=1)

        self.idx_lbl = ctk.CTkLabel(self, text=f"#{index:02d}", width=40, font=ctk.CTkFont(size=12))
        self.idx_lbl.grid(row=0, column=0, padx=(8, 2), pady=2)

        name_short = name if len(name) <= 14 else name[:13] + "…"
        self.name_lbl = ctk.CTkLabel(self, text=name_short, anchor="w", font=ctk.CTkFont(size=12))
        self.name_lbl.grid(row=0, column=1, sticky="ew", padx=2, pady=2)

        self.status_lbl = ctk.CTkLabel(self, text=status, width=30, font=ctk.CTkFont(size=12))
        self.status_lbl.grid(row=0, column=2, padx=(2, 8), pady=2)

        self._bind_click(self)
        self._bind_click(self.idx_lbl)
        self._bind_click(self.name_lbl)
        self._bind_click(self.status_lbl)

        self._update_style()

    def _bind_click(self, widget):
        for ev in ("<Button-1>", "<Enter>", "<Leave>"):
            widget.bind(ev, self._on_event)
        widget.configure(cursor="hand2")

    def _on_event(self, event):
        if event.type == tk.EventType.Enter:
            if not self.is_selected:
                self.configure(fg_color=("#E3F2FD", "#1A3A5C"))
        elif event.type == tk.EventType.Leave:
            if not self.is_selected:
                self.configure(fg_color="transparent")
        elif event.type == tk.EventType.ButtonPress:
            if self.command:
                self.command(self.index)

    def set_selected(self, selected):
        self.is_selected = selected
        self._update_style()

    def set_status(self, status):
        self.status_lbl.configure(text=status)

    def _update_style(self):
        if self.is_selected:
            self.configure(fg_color=("#1565C0", "#1565C0"))
            for lbl in (self.idx_lbl, self.name_lbl, self.status_lbl):
                lbl.configure(text_color=("#FFFFFF", "#FFFFFF"))
        else:
            self.configure(fg_color="transparent")
            self.idx_lbl.configure(text_color=("#1565C0", "#90CAF9"))
            self.name_lbl.configure(text_color=("#212121", "#E0E0E0"))
            self.status_lbl.configure(text_color=("#757575", "#9E9E9E"))

# ===================== 主应用 =====================
class App(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title("物料图片匹配系统")
        self.geometry("1300x800")
        self.minsize(1100, 700)

        # 数据
        self.owner_data = []
        self.images = []
        self.matcher = None
        self.match_results = []
        self.image_dir_path = None
        self.excel_path = None
        self.threshold = DEFAULT_THRESHOLD
        self.top_k = DEFAULT_TOP_K

        # 预览状态
        self.pil_image = None
        self.current_img_path = None
        self.current_filename = ""
        self.current_score = 0
        self.rotation_angle = 0
        self.zoom_level = ZOOM_FIT  # -1 = fit to panel
        self.selected_row_idx = None
        self.material_rows = []
        self._ctk_img_ref = None
        self._user_selections = {}  # 用户手动选择的图片 {item_idx: img_path}

        # 构建UI
        self._build_ui()

    # ===================== UI 构建 =====================

    def _build_ui(self):
        self.container = ctk.CTkFrame(self)
        self.container.pack(fill="both", expand=True)

        self.welcome_frame = ctk.CTkFrame(self.container)
        self.main_frame = ctk.CTkFrame(self.container)

        self._build_welcome()
        self._build_main()

        self._show_welcome()

    def _build_welcome(self):
        f = self.welcome_frame
        # 蓝色渐变背景（用浅蓝色frame模拟）
        f.configure(fg_color="#E3F2FD")

        # 居中卡片
        card = ctk.CTkFrame(f, corner_radius=16, fg_color="#FFFFFF")
        card.place(relx=0.5, rely=0.46, anchor="center")

        # 标题
        ctk.CTkLabel(card, text="物料图片匹配系统", font=ctk.CTkFont(size=26, weight="bold"),
                      text_color="#1565C0").pack(pady=(30, 5))
        ctk.CTkLabel(card, text="选择图片文件夹和物料清单，自动匹配输出图片",
                      font=ctk.CTkFont(size=13), text_color="#757575").pack(pady=(0, 25))

        # 选择区域
        sel_frame = ctk.CTkFrame(card, corner_radius=12, fg_color="#F8F9FA")
        sel_frame.pack(fill="x", padx=30, pady=5)

        # 图片
        r1 = ctk.CTkFrame(sel_frame, fg_color="transparent")
        r1.pack(fill="x", pady=8)
        self.btn_img = ctk.CTkButton(r1, text="  📂  选择图片文件夹", command=self._select_image_folder,
                                       height=40, font=ctk.CTkFont(size=13), fg_color="#1565C0")
        self.btn_img.pack(side="left")
        self.img_path_lbl = ctk.CTkLabel(r1, text="未选择", text_color="#757575",
                                           font=ctk.CTkFont(size=12))
        self.img_path_lbl.pack(side="left", padx=(12, 5))
        self.img_cnt_lbl = ctk.CTkLabel(r1, text="", text_color="#1565C0")
        self.img_cnt_lbl.pack(side="left")

        # Excel
        r2 = ctk.CTkFrame(sel_frame, fg_color="transparent")
        r2.pack(fill="x", pady=8)
        self.btn_excel = ctk.CTkButton(r2, text="  📄  加载物料清单", command=self._select_excel_file,
                                         height=40, font=ctk.CTkFont(size=13), fg_color="#1565C0")
        self.btn_excel.pack(side="left")
        self.excel_path_lbl = ctk.CTkLabel(r2, text="未选择", text_color="#757575",
                                             font=ctk.CTkFont(size=12))
        self.excel_path_lbl.pack(side="left", padx=(12, 5))
        self.excel_cnt_lbl = ctk.CTkLabel(r2, text="", text_color="#1565C0")
        self.excel_cnt_lbl.pack(side="left")

        # 列名显示
        self.cols_lbl = ctk.CTkLabel(sel_frame, text="", font=ctk.CTkFont(size=12), text_color="#1565C0")
        self.cols_lbl.pack(fill="x", pady=(5, 10))

        # 按钮
        btn_f = ctk.CTkFrame(card, fg_color="transparent")
        btn_f.pack(pady=(20, 10))
        self.btn_start = ctk.CTkButton(btn_f, text="  开始匹配  ", command=self._go_main,
                                         height=42, font=ctk.CTkFont(size=14), state="disabled", width=160,
                                         fg_color="#1565C0")
        self.btn_start.pack(side="left", padx=5)
        ctk.CTkButton(btn_f, text="  下载模板  ", command=self._download_template,
                       height=42, font=ctk.CTkFont(size=14), width=140,
                       fg_color="#78909C").pack(side="left", padx=5)

        self.welcome_status = ctk.CTkLabel(card, text="请选择图片文件夹和物料清单",
                                            font=ctk.CTkFont(size=12), text_color="#757575")
        self.welcome_status.pack(pady=(10, 25))

        ctk.CTkLabel(card, text="物料清单需包含「物料名称」列，支持 Excel (.xlsx)",
                     font=ctk.CTkFont(size=11), text_color="#9E9E9E").pack(pady=(0, 20))

    def _build_main(self):
        f = self.main_frame
        f.grid_rowconfigure(1, weight=1)
        f.grid_columnconfigure(0, weight=1)

        # --- 工具栏 ---
        toolbar = ctk.CTkFrame(f, corner_radius=0, height=50, fg_color="#FFFFFF")
        toolbar.grid(row=0, column=0, sticky="ew", padx=0, pady=0)
        toolbar.grid_columnconfigure(3, weight=1)

        btn_fg = "#1565C0"
        ctk.CTkButton(toolbar, text="← 返回", command=self._show_welcome, width=60).grid(row=0, column=0, padx=(10, 3), pady=8)
        ctk.CTkLabel(toolbar, text="|", text_color="#BDBDBD").grid(row=0, column=1, padx=2)
        self.btn_match = ctk.CTkButton(toolbar, text="全量匹配", command=self._run_matching, state="disabled", fg_color=btn_fg)
        self.btn_match.grid(row=0, column=2, padx=3, pady=8)
        self.btn_export = ctk.CTkButton(toolbar, text="导出结果", command=self._export_results, state="disabled", fg_color=btn_fg)
        self.btn_export.grid(row=0, column=4, padx=3, pady=8)
        ctk.CTkButton(toolbar, text="设置", command=self._open_settings, fg_color="#78909C").grid(row=0, column=5, padx=3, pady=8)

        self.src_lbl = ctk.CTkLabel(toolbar, text="", text_color="#757575", font=ctk.CTkFont(size=11))
        self.src_lbl.grid(row=0, column=7, sticky="e", padx=10, pady=8)

        self.progress = ctk.CTkProgressBar(toolbar, width=140, progress_color="#1565C0")
        self.progress.grid(row=0, column=6, padx=5, pady=8)
        self.progress.set(0)
        self.prog_lbl = ctk.CTkLabel(toolbar, text="", font=ctk.CTkFont(size=11))
        self.prog_lbl.grid(row=0, column=8, padx=(0, 5), pady=8)

        # --- 三面板 ---
        paned = ctk.CTkFrame(f)
        paned.grid(row=1, column=0, sticky="nsew", padx=8, pady=(0, 8))
        paned.grid_columnconfigure((0, 1, 2), weight=0, minsize=180)
        paned.grid_columnconfigure(0, weight=1)
        paned.grid_columnconfigure(1, weight=2)
        paned.grid_columnconfigure(2, weight=2)
        paned.grid_rowconfigure(0, weight=1)

        # 左：物料清单
        left_frame = ctk.CTkFrame(paned, corner_radius=10, fg_color="#FFFFFF")
        left_frame.grid(row=0, column=0, sticky="nsew", padx=(0, 4))
        ctk.CTkLabel(left_frame, text="物料清单", font=ctk.CTkFont(size=14, weight="bold"),
                      text_color="#1565C0").pack(anchor="w", padx=12, pady=(12, 5))
        self.list_container = ctk.CTkScrollableFrame(left_frame, corner_radius=8, fg_color="#FAFAFA")
        self.list_container.pack(fill="both", expand=True, padx=8, pady=(0, 8))

        # 中：详情
        center_frame = ctk.CTkFrame(paned, corner_radius=10, fg_color="#FFFFFF")
        center_frame.grid(row=0, column=1, sticky="nsew", padx=4)
        self._build_detail_panel(center_frame)

        # 右：预览
        right_frame = ctk.CTkFrame(paned, corner_radius=10, fg_color="#FFFFFF")
        right_frame.grid(row=0, column=2, sticky="nsew", padx=(4, 0))
        self._build_preview_panel(right_frame)

        # --- 状态栏 ---
        self.status_var = ctk.StringVar(value="就绪")
        status_bar = ctk.CTkLabel(f, textvariable=self.status_var, anchor="w", fg_color="#E3F2FD",
                                   corner_radius=0, height=28, font=ctk.CTkFont(size=11))
        status_bar.grid(row=2, column=0, sticky="ew")

    def _build_detail_panel(self, parent):
        ctk.CTkLabel(parent, text="物料信息", font=ctk.CTkFont(size=14, weight="bold"), text_color="#1565C0").pack(anchor="w", padx=12, pady=(12, 5))

        info_f = ctk.CTkFrame(parent, corner_radius=8, fg_color="#F8F9FA")
        info_f.pack(fill="x", padx=10, pady=(0, 8))
        self.detail_labels = {}
        fields = [("名称:", "name"), ("型号:", "model"), ("品牌:", "brand"), ("参数:", "params")]
        for i, (label, key) in enumerate(fields):
            ctk.CTkLabel(info_f, text=label, width=55, anchor="e", text_color="#1565C0").grid(row=i, column=0, sticky="e", padx=(8, 4), pady=3)
            lbl = ctk.CTkLabel(info_f, text="-", anchor="w", wraplength=260, justify="left", text_color="#212121")
            lbl.grid(row=i, column=1, sticky="w", pady=3)
            self.detail_labels[key] = lbl

        # 第5行：匹配图片原始文件名（方便对比型号）
        sep_line = ctk.CTkFrame(info_f, height=1, fg_color="#BBDEFB")
        sep_line.grid(row=len(fields), column=0, columnspan=2, sticky="ew", padx=8, pady=4)
        ctk.CTkLabel(info_f, text="图片:", width=55, anchor="e", text_color="#1565C0").grid(row=len(fields)+1, column=0, sticky="e", padx=(8, 4), pady=3)
        self.detail_labels["match_file"] = ctk.CTkLabel(info_f, text="-", anchor="w", wraplength=260, justify="left", text_color="#E65100")
        self.detail_labels["match_file"].grid(row=len(fields)+1, column=1, sticky="w", pady=3)

        sep = ctk.CTkFrame(parent, height=2, fg_color="#BBDEFB")
        sep.pack(fill="x", padx=10, pady=5)

        ctk.CTkLabel(parent, text="匹配结果", font=ctk.CTkFont(size=14, weight="bold"), text_color="#1565C0").pack(anchor="w", padx=12, pady=(5, 2))
        self.match_status = ctk.CTkLabel(parent, text="等待匹配...", font=ctk.CTkFont(size=13))
        self.match_status.pack(anchor="w", padx=12, pady=2)
        self.match_score = ctk.CTkLabel(parent, text="", font=ctk.CTkFont(size=13))
        self.match_score.pack(anchor="w", padx=12, pady=(0, 5))
        self.match_files_f = ctk.CTkScrollableFrame(parent, corner_radius=8, fg_color="#F8F9FA")
        self.match_files_f.pack(fill="both", expand=True, padx=8, pady=(0, 8))

    def _build_preview_panel(self, parent):
        self.preview_parent = parent

        # 占位文字
        self.preview_placeholder = ctk.CTkLabel(parent, text="选择物料后\n显示匹配图片", font=ctk.CTkFont(size=14),
                                                  text_color="#9E9E9E", justify="center")
        self.preview_placeholder.pack(expand=True)

        # 视口容器（固定大小，裁剪图片）
        self.viewport_frame = ctk.CTkFrame(parent, corner_radius=6, fg_color="#FAFAFA")

        # 使用 tk.Canvas 做真正的裁剪容器
        self.preview_canvas = tk.Canvas(
            self.viewport_frame,
            highlightthickness=0,
            bg="#FAFAFA",
            cursor="crosshair",
        )
        self.preview_canvas.pack(fill="both", expand=True)
        self.preview_canvas.bind("<MouseWheel>", self._on_mousewheel)
        self._canvas_img_id = None

        # 图片说明
        self.preview_caption = ctk.CTkLabel(parent, text="", font=ctk.CTkFont(size=12),
                                              text_color="#757575", wraplength=400)

        # 底部工具条
        self.preview_toolbar = ctk.CTkFrame(parent, corner_radius=8, height=38, fg_color="#F8F9FA")
        self.preview_toolbar.pack_forget()

        btn_size = 32
        btn_fg = "#1565C0"
        ctk.CTkButton(self.preview_toolbar, text="−", width=btn_size, height=btn_size, fg_color=btn_fg,
                       command=lambda: self._zoom(ZOOM_STEP * -1)).pack(side="left", padx=2)
        self.zoom_lbl = ctk.CTkLabel(self.preview_toolbar, text="适合", width=45, font=ctk.CTkFont(size=12),
                                      text_color="#1565C0")
        self.zoom_lbl.pack(side="left", padx=2)
        ctk.CTkButton(self.preview_toolbar, text="+", width=btn_size, height=btn_size, fg_color=btn_fg,
                       command=lambda: self._zoom(ZOOM_STEP)).pack(side="left", padx=2)
        ctk.CTkButton(self.preview_toolbar, text="⟲", width=btn_size, height=btn_size,
                       command=self._fit_zoom, fg_color="#78909C").pack(side="left", padx=5)
        ctk.CTkButton(self.preview_toolbar, text="↺", width=btn_size, height=btn_size,
                       command=lambda: self._rotate(-90), fg_color="#78909C").pack(side="left", padx=2)
        ctk.CTkButton(self.preview_toolbar, text="↻", width=btn_size, height=btn_size,
                       command=lambda: self._rotate(90), fg_color="#78909C").pack(side="left", padx=2)
        ctk.CTkButton(self.preview_toolbar, text="🔍", width=btn_size, height=btn_size,
                       command=self._open_full_preview, fg_color="#1565C0").pack(side="left", padx=5)

    # ===================== 页面切换 =====================

    def _show_welcome(self):
        self.main_frame.pack_forget()
        self.welcome_frame.pack(fill="both", expand=True)

    def _go_main(self):
        if not self.images or not self.owner_data:
            return
        self.welcome_frame.pack_forget()
        self.main_frame.pack(fill="both", expand=True)

        src = f"📁 {os.path.basename(str(self.image_dir_path))} | 📄 {os.path.basename(str(self.excel_path))}"
        self.src_lbl.configure(text=src)

        self._refresh_list()
        self.matcher = MatcherEngine(self.images, self.threshold, self.top_k)
        self.btn_match.configure(state="normal")
        self._set_status(f"已加载: {len(self.owner_data)} 项物料, {len(self.images)} 张图片 | 点击「全量匹配」开始")

    # ===================== 物料列表 =====================

    def _refresh_list(self):
        for w in self.list_container.winfo_children():
            w.destroy()
        self.material_rows = []
        for i, item in enumerate(self.owner_data, 1):
            row = MaterialRow(self.list_container, i, item["物料名称"], "⏳", command=self._on_row_click)
            self.material_rows.append(row)

    def _on_row_click(self, index):
        self.selected_row_idx = index - 1
        # 高亮更新
        for i, row in enumerate(self.material_rows):
            row.set_selected(i == index - 1)

        item = self.owner_data[index - 1]
        self.detail_labels["name"].configure(text=item["物料名称"])
        self.detail_labels["model"].configure(text=item.get("型号", "-") or "-")
        self.detail_labels["brand"].configure(text=item.get("品牌", "-") or "-")
        self.detail_labels["params"].configure(text=item.get("参数", "-") or "-")

        for w in self.match_files_f.winfo_children():
            w.destroy()

        if self.match_results and index - 1 < len(self.match_results):
            self._display_result(self.match_results[index - 1])
        else:
            self.match_status.configure(text="尚未匹配")
            self.match_score.configure(text="")
            self._show_placeholder()

    def _display_result(self, result):
        matches = result["matches"]
        # 清除旧内容
        for w in self.match_files_f.winfo_children():
            w.destroy()

        if not matches:
            self.match_status.configure(text="❌ 未匹配到图片", text_color="red")
            self.match_score.configure(text="")
            ctk.CTkLabel(self.match_files_f, text="未找到匹配图片\n请尝试降低匹配阈值",
                         text_color="gray", justify="center").pack(expand=True, pady=20)
            self._show_placeholder()
            return

        # 确定当前选中的图片
        sel_idx = self._get_selected_match_idx(result)
        best_img, best_score = matches[sel_idx]
        self.match_status.configure(text="✅ 已匹配", text_color="#1565C0")
        self.match_score.configure(text=f"当前选中: {best_score} 分  |  点击下方卡片切换", text_color="#1565C0")

        # 提示文字
        hint = ctk.CTkLabel(self.match_files_f, text="候选图片（点击切换）",
                            font=ctk.CTkFont(size=11), text_color="#757575")
        hint.pack(anchor="w", padx=4, pady=(0, 6))

        # 缩略图卡片容器
        cards_frame = ctk.CTkFrame(self.match_files_f, fg_color="transparent")
        cards_frame.pack(fill="x", padx=2)

        # 存储卡片组件引用（防GC）
        self._thumb_refs = []

        for idx, (img, score) in enumerate(matches):
            is_active = (idx == sel_idx)
            card = self._build_thumb_card(cards_frame, img, score, idx, is_active, result)
            card.pack(side="left", padx=4, pady=4)
            self._thumb_refs.append(card)

        # 显示选中图片
        self._show_image(best_img["path"], best_img["filename"], best_score)

    def _get_selected_match_idx(self, result):
        """获取当前应选中的匹配索引（用户手动选择优先）"""
        item_idx = None
        for i, r in enumerate(self.match_results):
            if r is result:
                item_idx = i
                break
        if item_idx is not None and hasattr(self, '_user_selections') and item_idx in self._user_selections:
            user_path = self._user_selections[item_idx]
            for j, (img, _) in enumerate(result["matches"]):
                if img["path"] == user_path:
                    return j
        return 0  # 默认选第一个

    def _build_thumb_card(self, parent, img_info, score, idx, is_active, result):
        """构建一张候选缩略图卡片"""
        card = ctk.CTkFrame(parent, corner_radius=8, width=85, height=110)
        card.pack_propagate(False)

        # 边框：选中时蓝色，未选时灰色
        border_color = "#1565C0" if is_active else "#E0E0E0"
        card.configure(fg_color=border_color)

        # 内衬白色背景
        inner = ctk.CTkFrame(card, corner_radius=6, fg_color="#FFFFFF")
        inner.pack(fill="both", expand=True, padx=2, pady=2)

        try:
            pil = Image.open(img_info["path"])
            pil.thumbnail(THUMB_CARD_SIZE, Image.LANCZOS)
            ctk_img = ctk.CTkImage(pil, size=pil.size)
        except Exception:
            ctk_img = None

        img_lbl = ctk.CTkLabel(inner, image=ctk_img, text="")
        img_lbl.pack(expand=True, pady=(4, 0))
        if ctk_img:
            self._ctk_img_ref = ctk_img  # 防止GC

        score_lbl = ctk.CTkLabel(inner, text=f"{score}分",
                                  font=ctk.CTkFont(size=10, weight="bold"),
                                  text_color="#1565C0" if is_active else "#757575")
        score_lbl.pack(pady=(1, 3))

        # 点击切换
        def on_click(e, i=idx, r=result):
            self._on_thumb_click(i, r)

        for w in (card, inner, img_lbl, score_lbl):
            w.bind("<Button-1>", on_click)
            w.configure(cursor="hand2")

        return card

    def _on_thumb_click(self, match_idx, result):
        """用户点击了某张候选缩略图"""
        # 找到这个 result 在 match_results 中的索引
        item_idx = None
        for i, r in enumerate(self.match_results):
            if r is result:
                item_idx = i
                break
        if item_idx is None:
            return

        img, score = result["matches"][match_idx]

        # 存储用户选择
        if not hasattr(self, '_user_selections'):
            self._user_selections = {}
        self._user_selections[item_idx] = img["path"]

        # 刷新当前选中项的显示
        self._display_result(result)

    # ===================== 图片预览（含缩放/旋转） =====================

    def _show_image(self, img_path, filename, score):
        try:
            self.pil_image = Image.open(img_path)
            self.current_img_path = img_path
            self.current_filename = filename
            self.current_score = score
            self.rotation_angle = 0
            self.zoom_level = ZOOM_FIT

            # 更新中间面板的匹配图片名
            self.detail_labels["match_file"].configure(text=filename)

            # 显示视口容器（填充分隔 + 说明 + 工具条），隐藏占位
            self.preview_placeholder.pack_forget()
            self.viewport_frame.pack(fill="both", expand=True, padx=10, pady=(10, 3))

            # 强制更新布局，确保 Canvas 有正确尺寸
            self.viewport_frame.update_idletasks()
            self.preview_canvas.update_idletasks()

            self._render_preview()

            self.preview_caption.configure(text=f"{filename}  |  {score} 分")
            self.preview_caption.pack(pady=(0, 3))

            self.preview_toolbar.pack(fill="x", padx=15, pady=(0, 10))
            self.zoom_lbl.configure(text="适合")

        except Exception as e:
            self._show_placeholder(f"无法加载图片:\n{e}")

    def _render_preview(self):
        if self.pil_image is None:
            return

        # 旋转
        img = self.pil_image.copy()
        if self.rotation_angle != 0:
            img = img.rotate(self.rotation_angle, expand=True, resample=Image.BICUBIC)

        # 获取 Canvas 实际可用尺寸
        cw = self.preview_canvas.winfo_width() - 4
        ch = self.preview_canvas.winfo_height() - 4
        if cw < 20 or ch < 20:
            cw, ch = 400, 300  # 保底尺寸

        iw, ih = img.size

        if self.zoom_level == ZOOM_FIT:
            ratio = min(cw / iw, ch / ih, 1.0)
        else:
            fit_ratio = min(cw / iw, ch / ih)
            ratio = fit_ratio * self.zoom_level

        new_w = max(50, int(iw * ratio))
        new_h = max(50, int(ih * ratio))
        img_r = img.resize((new_w, new_h), Image.LANCZOS)

        # 转为 PhotoImage (tkinter 原生格式) 用于 Canvas
        photo = ImageTk.PhotoImage(img_r)
        self._canvas_photo = photo  # 防GC

        # 居中绘制
        cx, cy = cw // 2, ch // 2
        if self._canvas_img_id is not None:
            self.preview_canvas.delete(self._canvas_img_id)
        self._canvas_img_id = self.preview_canvas.create_image(cx, cy, image=photo, anchor="center")

        # Canvas 区域标记（矩形边框）
        self.preview_canvas.configure(scrollregion=self.preview_canvas.bbox("all"))

        # 更新缩放显示
        if self.zoom_level == ZOOM_FIT:
            self.zoom_lbl.configure(text="适合")
        else:
            self.zoom_lbl.configure(text=f"{int(self.zoom_level * 100)}%")

    def _zoom(self, delta):
        if self.pil_image is None:
            return
        if self.zoom_level == ZOOM_FIT:
            self.zoom_level = 1.0
        self.zoom_level = max(ZOOM_MIN, min(ZOOM_MAX, self.zoom_level + delta))
        self._render_preview()

    def _fit_zoom(self):
        if self.pil_image is None:
            return
        self.zoom_level = ZOOM_FIT
        self._render_preview()

    def _on_mousewheel(self, event):
        delta = ZOOM_STEP if event.delta > 0 else -ZOOM_STEP
        self._zoom(delta)

    def _rotate(self, angle):
        if self.pil_image is None:
            return
        self.rotation_angle = (self.rotation_angle + angle) % 360
        self._render_preview()

    def _open_full_preview(self):
        if self.current_img_path:
            ImagePreviewWindow(self, self.current_img_path, self.current_filename)

    def _show_placeholder(self, text=None):
        self.viewport_frame.pack_forget()
        self.preview_caption.pack_forget()
        self.preview_toolbar.pack_forget()
        self.preview_placeholder.configure(text=text or "选择物料后\n显示匹配图片")
        self.preview_placeholder.pack(expand=True)
        self.detail_labels["match_file"].configure(text="-")
        self.pil_image = None
        self._canvas_img_id = None

    # ===================== 数据加载 =====================

    def _select_image_folder(self):
        d = filedialog.askdirectory(title="选择图片文件夹")
        if not d:
            return
        try:
            self.images = DataLoader.load_images(d)
            self.image_dir_path = Path(d)
            self.img_path_lbl.configure(text=str(self.image_dir_path))
            self.img_cnt_lbl.configure(text=f"({len(self.images)} 张)")
            self._check_ready()
        except Exception as e:
            messagebox.showerror("加载失败", str(e))

    def _select_excel_file(self):
        fp = filedialog.askopenfilename(title="选择物料清单", filetypes=[("Excel 文件", "*.xlsx *.xls"), ("所有文件", "*.*")])
        if not fp:
            return
        try:
            self.owner_data, cols = DataLoader.load_excel(fp)
            self.excel_path = Path(fp)
            self.excel_path_lbl.configure(text=str(self.excel_path))
            self.excel_cnt_lbl.configure(text=f"({len(self.owner_data)} 项)")
            txt = "识别列: " + "  ".join(f"「{k}」{'✓' if v is not None else '✗'}" for k, v in cols.items())
            self.cols_lbl.configure(text=txt)
            self._check_ready()
        except ValueError as e:
            messagebox.showerror("列名识别失败", str(e))
        except Exception as e:
            messagebox.showerror("加载失败", str(e))

    def _check_ready(self):
        if self.images and self.owner_data:
            self.btn_start.configure(state="normal", text="🚀  开始匹配")
            self.welcome_status.configure(text=f"✅ 就绪！{len(self.owner_data)} 项物料, {len(self.images)} 张图片",
                                           text_color="green")
        else:
            self.btn_start.configure(state="disabled")

    def _download_template(self):
        sp = filedialog.asksaveasfilename(title="保存模板", defaultextension=".xlsx",
                                           filetypes=[("Excel 文件", "*.xlsx")], initialfile="物料清单模板.xlsx")
        if not sp:
            return
        try:
            generate_template_excel(sp)
            messagebox.showinfo("模板已生成", f"✅ 模板已保存到:\n{sp}\n\n填写数据后，在应用中选择此文件即可")
        except Exception as e:
            messagebox.showerror("生成失败", str(e))

    # ===================== 匹配逻辑 =====================

    def _run_matching(self):
        if not self.owner_data or not self.matcher:
            return
        self.btn_match.configure(state="disabled")
        self.btn_export.configure(state="disabled")
        self.match_results = []

        self._prog_queue = queue.Queue()

        def cb(c, t, n):
            self._prog_queue.put((c, t, n))

        def worker():
            try:
                r = self.matcher.match_all(self.owner_data, cb)
                self._prog_queue.put(("DONE", r))
            except Exception as e:
                self._prog_queue.put(("ERR", str(e)))

        threading.Thread(target=worker, daemon=True).start()
        self._poll_progress()

    def _poll_progress(self):
        try:
            while True:
                msg = self._prog_queue.get_nowait()
                if msg[0] == "DONE":
                    self.match_results = msg[1]
                    self._on_match_done()
                    return
                elif msg[0] == "ERR":
                    self._set_status(f"匹配失败: {msg[1]}")
                    messagebox.showerror("匹配失败", msg[1])
                    self.btn_match.configure(state="normal")
                    return
                else:
                    c, t, n = msg
                    self.progress.set(c / t)
                    self.prog_lbl.configure(text=f"{c}/{t}")
                    self._set_status(f"匹配中 ({c}/{t}): {n[:18]}…")
        except queue.Empty:
            pass
        self.after(50, self._poll_progress)

    def _on_match_done(self):
        self.progress.set(1)
        mc = sum(1 for r in self.match_results if r["matched"])
        self._set_status(f"匹配完成: {mc}/{len(self.match_results)} ({mc/len(self.match_results)*100:.0f}%)")
        self.btn_match.configure(state="normal")
        self.btn_export.configure(state="normal")

        for i, res in enumerate(self.match_results):
            if i < len(self.material_rows):
                self.material_rows[i].set_status("✅" if res["matched"] else "❌")

        if self.selected_row_idx is not None and self.selected_row_idx < len(self.match_results):
            self._on_row_click(self.selected_row_idx + 1)

        messagebox.showinfo("匹配完成", f"✅ 全量匹配完成！\n\n总物料: {len(self.match_results)}\n匹配成功: {mc}")

    # ===================== 导出 =====================

    def _export_results(self):
        if not self.match_results:
            messagebox.showwarning("无结果", "请先执行全量匹配")
            return
        default = str(OUTPUT_DIR)
        if self.excel_path:
            default = str(OUTPUT_DIR / f"匹配结果_{self.excel_path.stem}")
        dp = filedialog.askdirectory(title="选择导出目录", initialdir=default)
        if not dp:
            return

        self._set_status("正在导出...")
        self.btn_export.configure(state="disabled")
        try:
            src = str(self.image_dir_path) if self.image_dir_path else ""
            st = Exporter.export_all(self.match_results, dp, self.threshold, src,
                                       user_selections=self._user_selections)
            self._set_status(f"导出成功: {st['matched']}/{st['total']} 项, {st['images_copied']} 张图片")
            self.btn_export.configure(state="normal")
            messagebox.showinfo("导出完成",
                f"✅ 导出成功！\n路径: {dp}\n"
                f"📊 Excel + 🌐 HTML + 🖼️ {st['images_copied']} 张图片\n"
                f"命名格式: 01-物料名称-型号.jpg")
            try:
                import subprocess
                if sys.platform == "darwin":
                    subprocess.run(["open", dp])
                elif sys.platform == "win32":
                    os.startfile(dp)
                else:
                    subprocess.run(["xdg-open", dp])
            except Exception:
                pass
        except Exception as e:
            self._set_status(f"导出失败: {e}")
            self.btn_export.configure(state="normal")
            messagebox.showerror("导出失败", str(e))

    # ===================== 设置 =====================

    def _open_settings(self):
        dlg = ctk.CTkToplevel(self)
        dlg.title("匹配设置")
        dlg.geometry("350x220")
        dlg.resizable(False, False)
        dlg.transient(self)
        dlg.grab_set()

        f = ctk.CTkFrame(dlg, corner_radius=12)
        f.pack(fill="both", expand=True, padx=20, pady=20)

        ctk.CTkLabel(f, text="匹配阈值 (10-95):").grid(row=0, column=0, sticky="w", pady=5)
        th_var = ctk.IntVar(value=self.threshold)
        ctk.CTkEntry(f, textvariable=th_var, width=80).grid(row=0, column=1, pady=5, sticky="w")

        ctk.CTkLabel(f, text="每项展示图片数:").grid(row=1, column=0, sticky="w", pady=5)
        tk_var = ctk.IntVar(value=self.top_k)
        ctk.CTkEntry(f, textvariable=tk_var, width=80).grid(row=1, column=1, pady=5, sticky="w")

        ctk.CTkLabel(f, text="阈值越高匹配越精确", text_color="gray", font=ctk.CTkFont(size=11)).grid(
            row=2, column=0, columnspan=2, pady=5)

        def save():
            try:
                self.threshold = max(10, min(95, th_var.get()))
                self.top_k = max(1, min(5, tk_var.get()))
                if self.matcher:
                    self.matcher.set_params(self.threshold, self.top_k)
                self._set_status(f"设置更新: 阈值={self.threshold}, 展示={self.top_k} 张")
                self.match_results = []
                for r in self.material_rows:
                    r.set_status("⏳")
                self.btn_export.configure(state="disabled")
                dlg.destroy()
            except Exception:
                pass

        ctk.CTkButton(f, text="保存", command=save).grid(row=3, column=0, pady=15)
        ctk.CTkButton(f, text="取消", command=dlg.destroy).grid(row=3, column=1, pady=15)

    def _set_status(self, text):
        self.status_var.set(text)
        self.update_idletasks()


# ===================== 入口 =====================
def main():
    try:
        import openpyxl  # noqa
    except ImportError:
        messagebox.showerror("缺少依赖", "请安装: pip install openpyxl")
        return
    try:
        import jieba  # noqa
    except ImportError:
        messagebox.showwarning("缺少 jieba", "建议安装: pip install jieba")

    app = App()
    app.mainloop()


if __name__ == "__main__":
    main()
